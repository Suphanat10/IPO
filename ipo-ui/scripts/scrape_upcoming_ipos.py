#!/usr/bin/env python3
"""
Scrape upcoming IPOs from SET + mai → upsert into PostgreSQL DB.

Data sources:
  - SET API: https://www.set.or.th/api/set/ipo/upcoming (SET + mai)
  - SEC กลต.: filing documents at market.sec.or.th
    • SUBSCRIPTION_AND_UNDERWRITION  → gross_proceeds, total_expense
    • SECURITIES_OFFERING            → offered_shares, offered_ratio_pct, existing_shares_pct
    • STRUCTURE / EXECUTIVE_INFO     → executive_total_pct
    • Financial Statements (Excel)   → BS & PL data

Usage:
  cd ipo-ui
  python scripts/scrape_upcoming_ipos.py              # full run
  python scripts/scrape_upcoming_ipos.py --dry-run    # fetch + display, no DB writes
  python scripts/scrape_upcoming_ipos.py --skip-sec-docs  # skip deep SEC document scraping

Requires:
  pip install curl_cffi pandas psycopg2-binary python-dotenv openpyxl
"""

import io
import sys
import os
import re
import json
import hashlib
import logging
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
from curl_cffi import requests as cffi_requests
from dotenv import load_dotenv
import psycopg2
import psycopg2.extras

# Concurrency for SEC scraping. Tune via env if SET/SEC rate-limits.
SEC_WORKERS = int(os.environ.get("SCRAPER_SEC_WORKERS", "3"))
SEC_DOC_WORKERS = int(os.environ.get("SCRAPER_SEC_DOC_WORKERS", "3"))
SET_TIMEOUT = int(os.environ.get("SCRAPER_SET_TIMEOUT", "20"))
SEC_PAGE_TIMEOUT = int(os.environ.get("SCRAPER_SEC_PAGE_TIMEOUT", "30"))
SEC_DOC_TIMEOUT = int(os.environ.get("SCRAPER_SEC_DOC_TIMEOUT", "45"))
SEC_FS_TIMEOUT = int(os.environ.get("SCRAPER_SEC_FS_TIMEOUT", "120"))
SEC_DOC_RETRIES = int(os.environ.get("SCRAPER_SEC_DOC_RETRIES", "1"))
SEC_RETRY_SLEEP_SECONDS = float(os.environ.get("SCRAPER_SEC_RETRY_SLEEP_SECONDS", "0.5"))
SEC_DOC_CACHE_ENABLED = os.environ.get("SCRAPER_SEC_DOC_CACHE", "1").lower() not in ("0", "false", "no")
SEC_DOC_CACHE_TTL_SECONDS = int(float(os.environ.get("SCRAPER_SEC_DOC_CACHE_TTL_HOURS", "168")) * 3600)
SEC_PAGE_CACHE_TTL_SECONDS = int(float(os.environ.get("SCRAPER_SEC_PAGE_CACHE_TTL_MINUTES", "30")) * 60)
SEC_DOC_PARSER_VERSION = "financials-v4"

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent
SEC_DOC_CACHE_DIR = SCRIPT_DIR / "output" / ".cache" / "sec-docs"
SEC_PAGE_CACHE_DIR = SCRIPT_DIR / "output" / ".cache" / "sec-pages"
load_dotenv(ROOT_DIR / ".env.local")
load_dotenv(ROOT_DIR / ".env")

DRY_RUN = "--dry-run" in sys.argv
SKIP_SEC_DOCS = "--skip-sec-docs" in sys.argv

# Optional: associate this run with a scrape_runs row in DB.
# Usage: --run-id <uuid>
RUN_ID: str | None = None
LOG_FILE: str | None = None
for i, arg in enumerate(sys.argv):
    if arg == "--run-id" and i + 1 < len(sys.argv):
        RUN_ID = sys.argv[i + 1]
    elif arg == "--log-file" and i + 1 < len(sys.argv):
        LOG_FILE = sys.argv[i + 1]

DATABASE_URL = os.environ.get("DATABASE_URL", "")
PG_HOST = os.environ.get("POSTGRES_HOST", "")
PG_PORT = int(os.environ.get("POSTGRES_PORT", "5432"))
PG_DB = os.environ.get("POSTGRES_DB", "")
PG_USER = os.environ.get("POSTGRES_USER", "")
PG_PASSWORD = os.environ.get("POSTGRES_PASSWORD", "")

SET_BASE = "https://www.set.or.th"
SET_PAGE = f"{SET_BASE}/th/listing/ipo/upcoming-ipo/set"
SET_API = f"{SET_BASE}/api/set/ipo/upcoming"

SEC_FILING_DETAIL = "https://market.sec.or.th/public/ipos/IPOSEQ01.aspx"

HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "th-TH,th;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": SET_PAGE,
}

SEC_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "th-TH,th;q=0.9,en-US;q=0.8,en;q=0.7",
    "Content-Type": "application/x-www-form-urlencoded",
}

SEC_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
)

_log_handlers: list[logging.Handler] = [logging.StreamHandler()]
if LOG_FILE:
    _log_file_path = Path(LOG_FILE) if os.path.isabs(LOG_FILE) else ROOT_DIR / LOG_FILE
    _log_file_path.parent.mkdir(parents=True, exist_ok=True)
    _log_handlers.append(logging.FileHandler(str(_log_file_path), encoding="utf-8"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=_log_handlers,
)
log = logging.getLogger(__name__)

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


def _json_cache_key(*parts: str) -> str:
    """Return a stable short key for versioned external resources."""
    payload = "\n".join(parts).encode("utf-8")
    return hashlib.sha1(payload).hexdigest()[:20]


def _load_json_cache(path: Path) -> dict | None:
    if not SEC_DOC_CACHE_ENABLED or not path.exists():
        return None
    if SEC_DOC_CACHE_TTL_SECONDS > 0 and time.time() - path.stat().st_mtime > SEC_DOC_CACHE_TTL_SECONDS:
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except Exception as e:
        log.debug("Cache read failed for %s: %s", path, e)
        return None


def _save_json_cache(path: Path, data: dict) -> None:
    if not SEC_DOC_CACHE_ENABLED:
        return
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(f"{path.suffix}.tmp")
        with tmp_path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, default=str)
        os.replace(tmp_path, path)
    except Exception as e:
        log.debug("Cache write failed for %s: %s", path, e)


def _sec_doc_cache_path(trans_id: str, urls: dict[str, str | None]) -> Path:
    version_parts = [SEC_DOC_PARSER_VERSION, trans_id]
    version_parts.extend(f"{name}={url or ''}" for name, url in sorted(urls.items()))
    return SEC_DOC_CACHE_DIR / f"{trans_id}_{_json_cache_key(*version_parts)}.json"


def _load_text_cache(path: Path, ttl_seconds: int, allow_stale: bool = False) -> str | None:
    if not SEC_DOC_CACHE_ENABLED or not path.exists():
        return None
    if not allow_stale and ttl_seconds > 0 and time.time() - path.stat().st_mtime > ttl_seconds:
        return None
    try:
        return path.read_text(encoding="utf-8")
    except Exception as e:
        log.debug("Cache read failed for %s: %s", path, e)
        return None


def _save_text_cache(path: Path, text: str) -> None:
    if not SEC_DOC_CACHE_ENABLED:
        return
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(f"{path.suffix}.tmp")
        tmp_path.write_text(text, encoding="utf-8")
        os.replace(tmp_path, path)
    except Exception as e:
        log.debug("Cache write failed for %s: %s", path, e)


def _sec_page_cache_path(trans_id: str) -> Path:
    return SEC_PAGE_CACHE_DIR / f"{trans_id}.html"


def _sec_request_headers() -> dict[str, str]:
    headers = dict(SEC_HEADERS)
    headers["User-Agent"] = SEC_USER_AGENT
    return headers


def _is_sec_rejection_text(text: str | None) -> bool:
    if not text:
        return False
    lowered = text.lower()
    return "request rejected" in lowered or "requested url was rejected" in lowered


def _is_valid_sec_page_html(text: str | None) -> bool:
    if not text or _is_sec_rejection_text(text):
        return False
    return "RadGrid1" in text or "IPOSGetFile.aspx" in text or "ContentPlaceHolder1_RadGrid1" in text


def _decode_http_content(content: bytes, headers: dict | None = None) -> str:
    charset = "utf-8"
    content_type = ""
    if headers:
        content_type = str(headers.get("Content-Type") or headers.get("content-type") or "")
    charset_match = re.search(r"charset=([\w\-]+)", content_type, re.IGNORECASE)
    if charset_match:
        charset = charset_match.group(1)
    return content.decode(charset, errors="replace")


def _urllib_get(url: str, timeout: int) -> SimpleNamespace | None:
    """Fallback SEC fetcher. Some SEC endpoints reset/reject curl_cffi but accept urllib."""
    try:
        req = Request(url, headers=_sec_request_headers())
        with urlopen(req, timeout=timeout) as resp:
            content = resp.read()
            headers = dict(resp.headers.items())
            return SimpleNamespace(
                status_code=resp.getcode(),
                content=content,
                text=_decode_http_content(content, headers),
                headers=headers,
                url=resp.geturl(),
            )
    except (HTTPError, URLError, TimeoutError, OSError) as e:
        log.debug("urllib SEC fetch failed for %s: %s", url, e)
        return None


def _response_looks_rejected(resp) -> bool:
    content = getattr(resp, "content", b"") or b""
    if len(content) > 4096 or content[:2] == b"PK":
        return False
    text = getattr(resp, "text", None)
    if text is None:
        text = content.decode("utf-8", errors="ignore")
    return _is_sec_rejection_text(text)


def _fetch_sec_page_html(session: cffi_requests.Session, filing_url: str) -> str | None:
    trans_id = extract_trans_id(filing_url)
    cache_path = _sec_page_cache_path(trans_id) if trans_id else None

    if cache_path:
        cached = _load_text_cache(cache_path, SEC_PAGE_CACHE_TTL_SECONDS)
        if cached and _is_valid_sec_page_html(cached):
            log.info("  SEC page: cache hit for TransID=%s", trans_id)
            return cached
        if cached:
            log.warning("  SEC page: ignoring invalid cache for TransID=%s", trans_id)

    try:
        resp = session.get(filing_url, headers=_sec_request_headers(), timeout=SEC_PAGE_TIMEOUT)
        if resp.status_code == 200 and _is_valid_sec_page_html(resp.text):
            if cache_path:
                _save_text_cache(cache_path, resp.text)
            return resp.text
        log.warning("SEC page %s returned invalid response (%s)", filing_url, getattr(resp, "status_code", "?"))
    except Exception as e:
        log.warning("SEC page fetch failed for %s: %s", filing_url, e)

    fallback = _urllib_get(filing_url, SEC_PAGE_TIMEOUT)
    if fallback and fallback.status_code == 200 and _is_valid_sec_page_html(fallback.text):
        if cache_path:
            _save_text_cache(cache_path, fallback.text)
        return fallback.text

    if cache_path:
        stale = _load_text_cache(cache_path, SEC_PAGE_CACHE_TTL_SECONDS, allow_stale=True)
        if stale and _is_valid_sec_page_html(stale):
            log.warning("  SEC page: using stale cache for TransID=%s", trans_id)
            return stale

    return None


# ---------------------------------------------------------------------------
# Utility: parse Thai-formatted numbers
# ---------------------------------------------------------------------------

def _parse_thai_number(text: str) -> float | None:
    """Parse a number from Thai text: '1,234,567.89' or '1234567.89' → float."""
    if not text:
        return None
    cleaned = re.sub(r"[^\d.,\-]", "", text.strip())
    cleaned = cleaned.replace(",", "")
    if not cleaned or cleaned in (".", "-", ""):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_thai_pct(text: str) -> float | None:
    """Parse a percentage like '25.50%' or '25.50' → 25.50."""
    if not text:
        return None
    cleaned = text.replace("%", "").replace("ร้อยละ", "").strip()
    return _parse_thai_number(cleaned)


# ---------------------------------------------------------------------------
# SET scraper
# ---------------------------------------------------------------------------

def fetch_upcoming_ipos() -> list[dict]:
    """Fetch upcoming IPOs from SET API (both SET and mai markets) in parallel."""
    session = cffi_requests.Session(impersonate="chrome")

    # Visit page first to establish cookies
    session.get(SET_PAGE, headers=HEADERS, timeout=SET_TIMEOUT)

    def _fetch_market(market_type: str) -> list[dict]:
        resp = session.get(
            SET_API,
            params={"type": market_type, "lang": "th"},
            headers=HEADERS,
            timeout=SET_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        items = data.get("data", []) if isinstance(data, dict) else data
        log.info("SET API (%s): %d IPOs", market_type, len(items))
        return items

    all_ipos: list[dict] = []
    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = {pool.submit(_fetch_market, m): m for m in ("SET", "mai")}
        for fut in as_completed(futures):
            try:
                all_ipos.extend(fut.result())
            except Exception as e:
                log.warning("SET API (%s) failed: %s", futures[fut], e)

    return all_ipos


# ---------------------------------------------------------------------------
# SEC document index: parse IPOSEQ01 RadGrid for direct download URLs
# ---------------------------------------------------------------------------

SEC_GETFILE_BASE = "https://market.sec.or.th/public/ipos/IPOSGetFile.aspx"


def _parse_filing_index(html: str) -> list[dict]:
    """
    Parse the RadGrid on IPOSEQ01.aspx to extract section titles and their
    direct download URLs (IPOSGetFile.aspx?TransID=X&TransFileSeq=Y).
    Returns list of { title, url, seq }.
    """
    sections: list[dict] = []
    row_pattern = re.compile(
        r'<tr[^>]*id="ctl00_ContentPlaceHolder1_RadGrid1_ctl00__\d+"[^>]*>(.*?)</tr>',
        re.DOTALL,
    )
    for tr_match in row_pattern.finditer(html):
        tr_html = tr_match.group(1)

        # Extract section title from first <td>
        first_td = re.search(r"<td[^>]*>(.*?)</td>", tr_html, re.DOTALL)
        if not first_td:
            continue
        title = re.sub(r"<[^>]+>", "", first_td.group(1))
        title = title.replace("&nbsp;", " ").replace("&nbsp", " ").replace("&amp;", "&")
        title = re.sub(r"\s+", " ", title).strip()
        if not title:
            continue

        # Extract direct download URLs from window.open() calls
        # Prefer the latest version (last URL in the row)
        file_urls = re.findall(
            r"window\.open\(&#39;(https://market\.sec\.or\.th/public/ipos/IPOSGetFile\.aspx\?[^']*?)&#39;\)",
            tr_html,
        )
        if not file_urls:
            continue

        # Use the last URL (latest filing version)
        best_url = file_urls[-1].replace("&amp;", "&")
        seq_match = re.search(r"TransFileSeq=(\d+)", best_url)
        seq = int(seq_match.group(1)) if seq_match else -1

        sections.append({
            "title": title,
            "url": best_url,
            "seq": seq,
        })

    return sections


def _find_section_url(sections: list[dict], *keywords: str) -> str | None:
    """Find the direct download URL for a section matching any keyword."""
    for sec in sections:
        if any(kw in sec["title"] for kw in keywords):
            return sec["url"]
    return None


def _find_latest_annual_fs_url(sections: list[dict]) -> str | None:
    """Find the download URL for the latest annual financial statement."""
    fs_sections: list[tuple[int, str]] = []
    for sec in sections:
        title = sec["title"]
        if "งบการเงิน" not in title:
            continue
        if re.search(r"ไตรมาส", title):
            continue
        year_match = re.search(r"(\d{4})", title)
        if year_match:
            year = int(year_match.group(1))
            fs_sections.append((year, sec["url"]))

    if not fs_sections:
        return None
    fs_sections.sort(key=lambda x: x[0], reverse=True)
    return fs_sections[0][1]


def _fetch_sec_document(
    session: cffi_requests.Session,
    url: str,
    timeout: int = SEC_DOC_TIMEOUT,
    retries: int = SEC_DOC_RETRIES,
) -> cffi_requests.Response | None:
    """Fetch a SEC filing document via direct IPOSGetFile URL, with retries."""
    for attempt in range(retries + 1):
        fallback = _urllib_get(url, timeout)
        if fallback and fallback.status_code == 200 and not _response_looks_rejected(fallback):
            return fallback

        try:
            resp = session.get(url, headers=_sec_request_headers(), timeout=timeout, allow_redirects=True)
            if resp.status_code == 200 and not _response_looks_rejected(resp):
                return resp
            log.warning("  SEC fetch returned invalid response (%s): %s", getattr(resp, "status_code", "?"), url)
        except Exception as e:
            if attempt < retries:
                log.warning("  SEC fetch attempt %d failed: %s — retrying...", attempt + 1, e)
                time.sleep(SEC_RETRY_SLEEP_SECONDS)
            else:
                log.warning("  SEC fetch failed after %d attempts: %s", retries + 1, e)
                return None


def _extract_docx_text(content: bytes) -> str:
    """Extract plain text from a DOCX file by parsing word/document.xml."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
        doc_xml = zf.read("word/document.xml").decode("utf-8")
        zf.close()
        texts = re.findall(r"<w:t[^>]*>([^<]+)</w:t>", doc_xml)
        return " ".join(texts)
    except Exception as e:
        log.warning("  DOCX text extraction failed: %s", e)
        return ""


def _is_xlsx_workbook(content: bytes) -> bool:
    """Check if bytes are an OOXML Excel workbook, even when the filename is .xls."""
    if not content or len(content) < 4 or content[:2] != b"PK":
        return False
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
        names = set(zf.namelist())
        zf.close()
        return "xl/workbook.xml" in names
    except Exception:
        return False


def _extract_xlsx_from_zip(content: bytes) -> bytes | None:
    """Extract the first Excel workbook from a ZIP archive."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
        excel_files = [
            f for f in zf.namelist()
            if f.lower().endswith((".xlsx", ".xlsm", ".xls"))
        ]
        for filename in excel_files:
            data = zf.read(filename)
            if _is_xlsx_workbook(data):
                log.info("  Extracted Excel: %s (%d bytes)", filename, len(data))
                zf.close()
                return data
        zf.close()
        return None
    except Exception as e:
        log.warning("  ZIP extraction failed: %s", e)
        return None


def _is_docx(content: bytes) -> bool:
    """Check if content is a DOCX file (ZIP with word/ directory)."""
    if not content or len(content) < 4:
        return False
    if content[:2] != b"PK":
        return False
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
        has_word = any("word/" in n for n in zf.namelist())
        zf.close()
        return has_word
    except Exception:
        return False


def _is_zip_with_xlsx(content: bytes) -> bool:
    """Check if content is a ZIP containing xlsx files."""
    if not content or len(content) < 4 or content[:2] != b"PK":
        return False
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
        has_xlsx = any(n.lower().endswith((".xlsx", ".xlsm", ".xls")) for n in zf.namelist())
        zf.close()
        return has_xlsx
    except Exception:
        return False


# ---------------------------------------------------------------------------
# SEC document parsers
# ---------------------------------------------------------------------------

_THAI_NUMBER_TOKEN = r"-?\d[\d\s,]*(?:\s*\.\s*\d+)?"


def _normalize_doc_text(text: str) -> str:
    text = (text or "").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _money_values_in_text(text: str) -> list[float]:
    values: list[float] = []
    for match in re.finditer(rf"({_THAI_NUMBER_TOKEN})\s*(ล้านบาท|บาท)", text):
        raw = match.group(1).replace(" ", "")
        val = _parse_thai_number(raw)
        if val is None:
            continue
        if match.group(2) == "ล้านบาท":
            val *= 1_000_000
        if abs(val) >= 1_000:
            values.append(val)
    return values


def _first_money_near_keywords(text: str, keywords: tuple[str, ...], window: int = 800) -> float | None:
    for kw in keywords:
        idx = text.find(kw)
        if idx == -1:
            continue
        nearby = text[idx:idx + window]
        values = _money_values_in_text(nearby)
        if values:
            return values[0]
    return None


def _first_pct(text: str) -> float | None:
    patterns = (
        rf"ร้อยละ\s*({_THAI_NUMBER_TOKEN})",
        rf"({_THAI_NUMBER_TOKEN})\s*%",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        val = _parse_thai_pct(match.group(1).replace(" ", ""))
        if val is not None and 0 <= val <= 100:
            return val
    return None


def _parse_subscription_report(text: str) -> dict:
    """
    Parse SUBSCRIPTION_AND_UNDERWRITION report for:
      - gross_proceeds (มูลค่าการเสนอขาย)
      - total_expense (ค่าใช้จ่ายในการเสนอขาย)
    """
    result: dict = {}
    text = _normalize_doc_text(text)

    gross = _first_money_near_keywords(
        text,
        ("ประมาณการจำนวนเงิน", "จำนวนเงินค่าหุ้น", "มูลค่าการเสนอขาย", "มูลค่ารวมของหุ้น"),
    )
    if gross is not None:
        result["gross_proceeds"] = gross
        log.info("    gross_proceeds = %s", gross)

    expense = _first_money_near_keywords(
        text,
        ("รวมค่าใช้จ่าย", "รวมค่าใช้จ่ายทั้งสิ้น", "ประมาณการค่าใช้จ่าย"),
    )
    if expense is not None:
        result["total_expense"] = expense
        log.info("    total_expense = %s", expense)

    return result


def _parse_securities_offering(text: str) -> dict:
    """
    Parse SECURITIES_OFFERING report for:
      - offered_shares (จำนวนหุ้นที่เสนอขาย)
      - offered_ratio_pct (สัดส่วนหุ้นที่เสนอขาย)
      - existing_shares_pct (สัดส่วนการขายหุ้นเดิม)
    """
    result: dict = {}
    text = _normalize_doc_text(text)

    # offered_shares
    for kw in ("ลักษณะสำคัญของหลักทรัพย์ที่เสนอขาย", "รายละเอียดของหลักทรัพย์ที่เสนอขาย",
               "จำนวนหุ้นที่เสนอขาย"):
        idx = text.find(kw)
        if idx == -1:
            continue
        nearby = text[idx:idx + 1000]
        # DOCX numbers may have spaces: "180 ,000,000 หุ้น"
        shares_match = re.search(rf"({_THAI_NUMBER_TOKEN})\s*หุ้น", nearby)
        if shares_match:
            raw = shares_match.group(1).replace(" ", "")
            val = _parse_thai_number(raw)
            if val and val > 100:
                result["offered_shares"] = int(val)
                log.info("    offered_shares = %s", int(val))
                break

    # offered_ratio_pct — search near the securities-offering detail first.
    ratio_text = text
    for section_kw in ("ลักษณะสำคัญของหลักทรัพย์ที่เสนอขาย", "รายละเอียดของหลักทรัพย์ที่เสนอขาย"):
        section_idx = text.find(section_kw)
        if section_idx != -1:
            ratio_text = text[section_idx:section_idx + 1500]
            break
    for kw in ("คิดเป็นร้อยละ", "ร้อยละ"):
        idx = ratio_text.find(kw)
        if idx == -1:
            continue
        nearby = ratio_text[idx:idx + 300]
        val = _first_pct(nearby)
        if val is not None and 0 < val <= 100:
            result["offered_ratio_pct"] = val
            log.info("    offered_ratio_pct = %s%%", val)
            break

    # existing_shares_pct
    existing_context_found = False
    for kw in ("หุ้นเดิม", "หุ้นสามัญเดิม", "ผู้ถือหุ้นเดิม"):
        matches = list(re.finditer(kw, text))
        for m in matches:
            nearby = text[m.start():m.start() + 500]
            if "เสนอขาย" in nearby or "จำหน่าย" in nearby:
                existing_context_found = True
            val = _first_pct(nearby)
            if val is not None and 0 < val <= 100:
                result["existing_shares_pct"] = val
                log.info("    existing_shares_pct = %s%%", val)
                break
        if "existing_shares_pct" in result:
            break
    if "existing_shares_pct" not in result and result.get("offered_shares") and not existing_context_found:
        result["existing_shares_pct"] = 0.0
        log.info("    existing_shares_pct = 0.0%%")

    return result


def _parse_shareholder_info(text: str) -> dict:
    """
    Parse STRUCTURE / EXECUTIVE_INFO for:
      - executive_total_pct (สัดส่วนการถือหุ้นของผู้บริหาร)

    The DOCX shareholder table typically has rows like:
      "รวมกลุ่มXXX 480,000,000 100.00 480,000,000 72.73"
    where the last number is the post-IPO percentage.
    """
    result: dict = {}
    text = _normalize_doc_text(text)

    # Find the shareholder section
    for kw in ("รายชื่อผู้ถือหุ้น", "โครงสร้างการถือหุ้น", "ผู้ถือหุ้นรายใหญ่"):
        idx = text.find(kw)
        if idx == -1:
            continue
        nearby = text[idx:idx + 5000]

        # Strategy 1: find "รวมกลุ่ม" row and extract the last percentage
        # Pattern: "รวมกลุ่ม... shares pct shares pct" — we want the last pct (post-IPO)
        group_match = re.search(
            r"รวมกลุ่ม[^\d]*"
            rf"[\d,\s]+\s+({_THAI_NUMBER_TOKEN})\s+"  # pct_before
            rf"[\d,\s]+\s+({_THAI_NUMBER_TOKEN})",     # pct_after (post-IPO)
            nearby,
        )
        if group_match:
            raw = group_match.group(2).replace(" ", "")
            val = _parse_thai_pct(raw)
            if val and 0 < val < 100:
                result["executive_total_pct"] = val
                log.info("    executive_total_pct = %s%% (from รวมกลุ่ม)", val)
                break

        # Strategy 2: shareholder tables often list shares/pct before and after IPO.
        # Use the first real shareholder row when there is no explicit "รวมกลุ่ม" row.
        row_pattern = re.compile(
            rf"(?P<name>(?:\d+\.\s*)?[^\d]{{3,120}}?)\s+"
            rf"(?P<shares_before>[\d,\s]+|-)\s+(?P<pct_before>{_THAI_NUMBER_TOKEN})\s+"
            rf"(?P<shares_after>[\d,\s]+|-)\s+(?P<pct_after>{_THAI_NUMBER_TOKEN})"
        )
        for row_match in row_pattern.finditer(nearby):
            name = re.sub(r"\s+", " ", row_match.group("name")).strip()
            if any(skip in name for skip in ("เสนอขาย", "ประชาชนทั่วไป", "IPO", "รวม")):
                continue
            val = _parse_thai_pct(row_match.group("pct_after").replace(" ", ""))
            if val is not None and 0 < val < 100:
                result["executive_total_pct"] = val
                log.info("    executive_total_pct = %s%% (from shareholder row)", val)
                break
        if "executive_total_pct" in result:
            break

        # Strategy 3: find explicit % symbols near ผู้บริหาร or กรรมการ
        exec_match = re.search(
            r"(?:ผู้บริหาร|กรรมการ|ผู้ถือหุ้นรายใหญ่).*?"
            rf"({_THAI_NUMBER_TOKEN})\s*%",
            nearby,
        )
        if exec_match:
            raw = exec_match.group(1).replace(" ", "")
            val = _parse_thai_pct(raw)
            if val and 0 < val <= 100:
                result["executive_total_pct"] = val
                log.info("    executive_total_pct = %s%% (from %%)", val)
                break

        # Strategy 4: find "รวม" row with numeric percentages (no % symbol)
        # In DOCX tables, data appears as: "รวม... 480,000,000 100.00 480,000,000 72.73"
        total_match = re.search(
            r"รวม(?:ทั้งหมด|กลุ่ม|ผู้ถือหุ้น)[^\d]*"
            rf"([\d][\d,\s]*[\d])\s+({_THAI_NUMBER_TOKEN})\s+"
            rf"([\d][\d,\s]*[\d])\s+({_THAI_NUMBER_TOKEN})",
            nearby,
        )
        if total_match:
            raw = total_match.group(4).replace(" ", "")
            val = _parse_thai_pct(raw)
            if val and 0 < val < 100:
                result["executive_total_pct"] = val
                log.info("    executive_total_pct = %s%% (from รวม row)", val)
                break

    return result


def _parse_financial_excel(content: bytes) -> dict:
    """
    Parse financial statements from Excel file.
    BS sheet → total_assets, total_liabilities, total_equity
    PL sheet → revenue_latest, revenue_prev, net_income_latest, net_income_prev
    """
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel financial parsing")
        return {}

    if content[:2] == b"PK" and not _is_xlsx_workbook(content):
        nested = _extract_xlsx_from_zip(content)
        if nested:
            return _parse_financial_excel(nested)

    result: dict = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        log.warning("Failed to open Excel: %s", e)
        return {}

    sheet_names = wb.sheetnames
    log.info("    Excel sheets: %s", sheet_names)

    def _pick_sheet(exact_name: str, fallback_pattern: str):
        for name in sheet_names:
            if name.strip().lower() == exact_name.lower():
                return wb[name]
        for name in sheet_names:
            if re.search(fallback_pattern, name, re.IGNORECASE):
                return wb[name]
        return None

    bs_sheets = [wb[name] for name in sheet_names if name.strip().lower() == "bs"]
    if not bs_sheets:
        bs_sheets = [
            wb[name] for name in sheet_names
            if re.search(r"^BS\b|balance|งบแสดงฐานะ|ฐานะการเงิน|สินทรัพย์|หนี้สิน", name, re.IGNORECASE)
        ]
    for bs_sheet in bs_sheets:
        for key, value in _parse_bs_sheet(bs_sheet).items():
            if key not in result:
                result[key] = value

    # --- Profit & Loss ---
    pl_sheet = _pick_sheet("PL", r"IS|CI|profit|loss|income|งบกำไร|กำไรขาดทุน|รายได้|เบ็ดเสร็จ")
    if pl_sheet:
        result.update(_parse_pl_sheet(pl_sheet))

    wb.close()
    return result


def _sheet_label_matches(cell_text: str, keyword_patterns: list[str], regex: bool = False) -> bool:
    cell_text = re.sub(r"\s+", " ", cell_text).strip()
    for pattern in keyword_patterns:
        if regex:
            if re.search(pattern, cell_text, re.IGNORECASE):
                return True
        elif pattern in cell_text:
            return True
    return False


def _find_value_in_sheet(
    sheet,
    keyword_patterns: list[str],
    col_offset: int = 1,
    regex: bool = False,
) -> float | None:
    """
    Search a sheet for a row matching any keyword pattern,
    then return the numeric value from the specified column offset.
    col_offset=1 means the column immediately right of the keyword cell.
    """
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value is None:
                continue
            cell_text = str(cell.value).strip()
            if _sheet_label_matches(cell_text, keyword_patterns, regex=regex):
                # Look for numeric values in the same row, to the right
                for offset in range(col_offset, 8):
                    try:
                        val_cell = row[cell.column - 1 + offset]
                        if val_cell.value is not None:
                            val = _parse_thai_number(str(val_cell.value))
                            if val is not None:
                                return val
                    except IndexError:
                        break
    return None


def _find_two_period_values(
    sheet,
    keyword_patterns: list[str],
    regex: bool = False,
) -> tuple[float | None, float | None]:
    """
    Find a row matching keywords and return (latest, previous) period values.
    Assumes latest period is in an earlier column (left = latest).
    """
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value is None:
                continue
            cell_text = str(cell.value).strip()
            if _sheet_label_matches(cell_text, keyword_patterns, regex=regex):
                values: list[float] = []
                for offset in range(1, 10):
                    try:
                        val_cell = row[cell.column - 1 + offset]
                        if val_cell.value is not None:
                            val = _parse_thai_number(str(val_cell.value))
                            if val is not None:
                                values.append(val)
                    except IndexError:
                        break
                if len(values) >= 2:
                    return values[0], values[1]
                elif len(values) == 1:
                    return values[0], None
    return None, None


def _parse_bs_sheet(sheet) -> dict:
    """Extract BS data: total_assets, total_liabilities, total_equity."""
    result: dict = {}

    val = _find_value_in_sheet(sheet, [r"^รวมสินทรัพย์$", r"^สินทรัพย์รวม$", r"^Total assets$"], regex=True)
    if val is not None:
        result["total_assets"] = val
        log.info("    total_assets = %s", val)

    val = _find_value_in_sheet(sheet, [r"^รวมหนี้สิน$", r"^หนี้สินรวม$", r"^Total liabilities$"], regex=True)
    if val is not None:
        result["total_liabilities"] = val
        log.info("    total_liabilities = %s", val)

    val = _find_value_in_sheet(sheet, [
        r"^รวมส่วนของผู้ถือหุ้น.*$", r"^รวมส่วนของเจ้าของ.*$",
        r"^ส่วนของผู้ถือหุ้นรวม$", r"^Total equity$",
        r"^Total shareholders.*equity$", r"^Total shareholders.*$",
    ], regex=True)
    if val is not None:
        result["total_equity"] = val
        log.info("    total_equity = %s", val)

    return result


def _parse_pl_sheet(sheet) -> dict:
    """Extract PL data: revenue (latest+prev), net_income (latest+prev)."""
    result: dict = {}

    rev_latest, rev_prev = _find_two_period_values(sheet, [
        r"^รวมรายได้$", r"^รายได้รวม$", r"^Total revenue$", r"^Total income$",
    ], regex=True)
    if rev_latest is not None:
        result["revenue_latest"] = rev_latest
        log.info("    revenue_latest = %s", rev_latest)
    if rev_prev is not None:
        result["revenue_prev"] = rev_prev
        log.info("    revenue_prev = %s", rev_prev)

    ni_latest, ni_prev = _find_two_period_values(sheet, [
        r"^กำไรสุทธิสำหรับปี$", r"^กำไรสำหรับปี$", r"^กำไร\(ขาดทุน\)สำหรับปี$",
        r"^Net income$", r"^Profit for the year$", r"^Net profit$",
    ], regex=True)
    if ni_latest is not None:
        result["net_income_latest"] = ni_latest
        log.info("    net_income_latest = %s", ni_latest)
    if ni_prev is not None:
        result["net_income_prev"] = ni_prev
        log.info("    net_income_prev = %s", ni_prev)

    return result


# ---------------------------------------------------------------------------
# SEC filing scraper (main entry point for each IPO)
# ---------------------------------------------------------------------------

def extract_trans_id(filing_url: str) -> str | None:
    """Extract TransID from SEC filing URL."""
    m = re.search(r"TransID=(\d+)", filing_url or "")
    return m.group(1) if m else None


def _parse_table_rows(html: str) -> list[list[str]]:
    """Parse HTML table rows into lists of cell text."""
    rows = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.DOTALL):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.DOTALL)
        if cells:
            cleaned = [
                re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", c))
                .replace("&nbsp;", "")
                .replace("&nbsp", "")
                .strip()
                for c in cells
            ]
            rows.append([c for c in cleaned if c and c != ":"])
    return rows


def scrape_sec_filing(session: cffi_requests.Session, filing_url: str) -> dict:
    """
    Scrape SEC filing page for:
      - FA person (ที่ปรึกษาทางการเงิน/ผู้ควบคุม)
      - Financial statement period count
      - Filing documents (gross_proceeds, total_expense, offered data, BS/PL)
    """
    if not filing_url:
        return {}

    try:
        text = _fetch_sec_page_html(session, filing_url)
        if not text:
            return {}
        result: dict = {}

        # --- FA person + FA company from "ที่ปรึกษาทางการเงิน/ผู้ควบคุม" row ---
        rows = _parse_table_rows(text)
        for row in rows:
            joined = " ".join(row)
            if "ที่ปรึกษาทางการเงิน" in joined:
                val = row[-1] if row else ""
                if "/" in val:
                    parts = [p.strip() for p in val.split("/", 1)]
                    result["fa_company_sec"] = parts[0]
                    fa_person = parts[1] if len(parts) > 1 else None
                    if fa_person and fa_person not in ("N.A.", "-", ""):
                        result["fa_person"] = fa_person
                        log.info("  SEC: FA person = %s", fa_person)
                    else:
                        log.info("  SEC: FA person not available")
                else:
                    result["fa_company_sec"] = val
                break

        # --- Financial statement periods ---
        fin_periods = re.findall(r"\[ส่วนที่ 3\] - งบการเงิน\s+([^<]+)", text)
        if fin_periods:
            result["financial_periods_available"] = [p.strip() for p in fin_periods]
            log.info("  SEC: found %d financial periods", len(fin_periods))

        # --- Deep document scraping (PostBack) ---
        if not SKIP_SEC_DOCS:
            trans_id = extract_trans_id(filing_url)
            if trans_id:
                doc_data = _scrape_sec_documents(session, trans_id, text)
                result["doc_financials"] = doc_data

        return result

    except Exception as e:
        log.warning("SEC scrape failed for %s: %s", filing_url, e)
        return {}


def _scrape_sec_documents(
    session: cffi_requests.Session,
    trans_id: str,
    filing_html: str,
) -> dict:
    """
    Scrape individual SEC filing documents via direct IPOSGetFile URLs.
    All documents are fetched in parallel for speed.
    """
    result: dict = {}

    sections = _parse_filing_index(filing_html)
    if not sections:
        log.warning("  SEC docs: no sections found for TransID=%s", trans_id)
        return result

    log.info("  SEC docs: found %d sections for TransID=%s", len(sections), trans_id)

    # Resolve all document URLs upfront. Direct SEC URLs include TransFileSeq,
    # so this cache invalidates automatically when a newer filing version appears.
    sub_url = _find_section_url(sections, "การจอง การจำหน่าย และการจัดสรร", "การจอง")
    sec_url = _find_section_url(sections, "รายละเอียดของหลักทรัพย์ที่เสนอขาย", "รายละเอียดของหลักทรัพย์")
    struct_url = _find_section_url(sections, "โครงสร้างและการดำเนินงาน")
    appendix_url = _find_section_url(sections, "รายละเอียดเกี่ยวกับกรรมการ", "เอกสารแนบ 1")
    fs_url = _find_latest_annual_fs_url(sections)

    selected_urls = {
        "subscription": sub_url,
        "securities": sec_url,
        "structure": struct_url,
        "appendix_executive": appendix_url,
        "financial_statements": fs_url,
    }
    cache_path = _sec_doc_cache_path(trans_id, selected_urls)
    cached = _load_json_cache(cache_path)
    if cached is not None:
        log.info("  SEC docs: cache hit for TransID=%s (%d fields)", trans_id, len(cached))
        return cached

    def _fetch_and_extract_text(doc_url: str) -> str:
        resp = _fetch_sec_document(session, doc_url)
        if not resp or resp.status_code != 200:
            return ""
        if _is_docx(resp.content):
            return _extract_docx_text(resp.content)
        text = re.sub(r"<[^>]+>", " ", resp.text)
        return re.sub(r"\s+", " ", text)

    # Fetch text-based documents in parallel
    text_tasks: list[tuple[str, str, callable]] = []
    if sub_url:
        text_tasks.append(("subscription", sub_url, _parse_subscription_report))
    if sec_url:
        text_tasks.append(("securities", sec_url, _parse_securities_offering))
    if struct_url:
        text_tasks.append(("structure", struct_url, _parse_shareholder_info))
    if appendix_url and appendix_url != struct_url:
        text_tasks.append(("appendix_executive", appendix_url, _parse_shareholder_info))

    def _fetch_text_doc(name: str, url: str, parser: callable) -> tuple[str, dict]:
        log.info("  SEC docs: fetching %s...", name)
        text = _fetch_and_extract_text(url)
        if not text:
            return f"{name}__fetch_failed", {}
        return name, parser(text)

    def _fetch_fs_doc() -> tuple[str, dict]:
        log.info("  SEC docs: fetching Financial Statements...")
        resp = _fetch_sec_document(session, fs_url, timeout=SEC_FS_TIMEOUT, retries=SEC_DOC_RETRIES)
        if not resp or resp.status_code != 200:
            return "fs__fetch_failed", {}
        log.info("  SEC docs: FS size = %d bytes", len(resp.content))
        if resp.content[:2] == b"PK":
            xlsx_data = _extract_xlsx_from_zip(resp.content)
            if xlsx_data:
                return "fs", _parse_financial_excel(xlsx_data)
        if _is_docx(resp.content):
            log.info("  SEC docs: FS is DOCX, skipping (need Excel)")
            return "fs", {}
        return "fs", _parse_financial_excel(resp.content)

    with ThreadPoolExecutor(max_workers=SEC_DOC_WORKERS) as pool:
        futures = []
        for name, url, parser in text_tasks:
            futures.append(pool.submit(_fetch_text_doc, name, url, parser))
        if fs_url:
            futures.append(pool.submit(_fetch_fs_doc))

        fetch_failed = False
        for fut in as_completed(futures):
            try:
                name, parsed = fut.result()
                if name.endswith("__fetch_failed"):
                    fetch_failed = True
                    continue
                for key, value in parsed.items():
                    if value is not None and key not in result:
                        result[key] = value
            except Exception as e:
                fetch_failed = True
                log.warning("  SEC docs: parallel fetch failed: %s", e)

    if result:
        log.info("  SEC docs: extracted %d fields from documents", len(result))
        if fetch_failed:
            log.warning("  SEC docs: not caching partial result for TransID=%s", trans_id)
        else:
            _save_json_cache(cache_path, result)
    else:
        log.info("  SEC docs: no financial data extracted from documents")

    return result


# ---------------------------------------------------------------------------
# Transform SET data → DB schema
# ---------------------------------------------------------------------------

def parse_offered_shares(no_of_ipo: str | None) -> int | None:
    """Extract number of offered shares from SET text like '612,451,687 หุ้น ...'"""
    if not no_of_ipo:
        return None
    m = re.match(r"([\d,]+)", no_of_ipo.replace(" ", ""))
    if m:
        try:
            return int(m.group(1).replace(",", ""))
        except ValueError:
            return None
    return None


def transform_ipo(raw: dict, sec_data: dict) -> dict:
    """Transform SET API response to match ipos + ipo_financials DB schema."""
    symbol = (raw.get("symbol") or "").strip().upper()
    if not symbol:
        return {}

    fa_companies = raw.get("financialAdvisors") or []
    lead_uw = raw.get("underwriters") or []

    first_trade = raw.get("firstTradeDate")
    listing_date = None
    if first_trade:
        try:
            listing_date = datetime.strptime(first_trade, "%Y-%m-%dT%H:%M:%S").date().isoformat()
        except (ValueError, TypeError):
            listing_date = first_trade

    fa_person = sec_data.get("fa_person")
    fa_persons_list = [fa_person] if fa_person else None

    ipo_price = raw.get("ipoPrice")
    par_value = raw.get("par")

    # Filing status: Approved / Effective / Submitted
    filing_status = raw.get("status")

    # Business description (ประเภทธุรกิจ)
    business_desc = raw.get("businessDescription")

    ipo_row = {
        "symbol": symbol,
        "company_name": raw.get("nameEn") or raw.get("nameTh"),
        "company_name_th": raw.get("nameTh"),
        "market": raw.get("market"),
        "industry": raw.get("industry"),
        "sector": raw.get("sector"),
        "status": "upcoming",
        "filing_status": filing_status,
        "business_description": business_desc,
        "listing_date": listing_date,
        "ipo_price": ipo_price,
        "par_value": par_value,
        "fa_persons": fa_persons_list,
        "fa_companies": [c.replace("\xa0", " ").strip() for c in fa_companies] if fa_companies else None,
        "lead_uw": [u.replace("\xa0", " ").strip() for u in lead_uw] if lead_uw else None,
        "source": "set_api_scraper",
    }

    # Build financials from SET API + SEC document scraping
    financials_row: dict = {}

    # From SET API
    offered_shares = parse_offered_shares(raw.get("noOfIPO"))
    if offered_shares:
        financials_row["offered_shares"] = offered_shares

    # From SEC document scraping (overrides SET data where available)
    doc_fin = sec_data.get("doc_financials") or {}
    for field in (
        "gross_proceeds", "total_expense",
        "offered_shares", "offered_ratio_pct", "existing_shares_pct",
        "executive_total_pct",
        "total_assets", "total_liabilities", "total_equity",
        "revenue_latest", "revenue_prev",
        "net_income_latest", "net_income_prev",
    ):
        val = doc_fin.get(field)
        if val is not None:
            financials_row[field] = val

    # SEC metadata (filing reference)
    sec_meta = {
        "filing_url": raw.get("filingUrl"),
        "sec_trans_id": extract_trans_id(raw.get("filingUrl")),
        "executive_summary_url": raw.get("executiveSummaryUrl"),
        "par_value": raw.get("par"),
        "pe_ratio": raw.get("pe"),
        "market_cap": raw.get("marketCap"),
        "issued_size": raw.get("issuedSize"),
        "financial_periods": sec_data.get("financial_periods_available"),
    }

    return {
        "ipo": {k: v for k, v in ipo_row.items() if v is not None},
        "financials": financials_row if financials_row else None,
        "sec_meta": {k: v for k, v in sec_meta.items() if v is not None},
    }


# ---------------------------------------------------------------------------
# PostgreSQL upsert
# ---------------------------------------------------------------------------

def _has_pg_config() -> bool:
    return bool(DATABASE_URL or (PG_HOST and PG_DB))


def _get_pg_conn():
    """Create a new PostgreSQL connection."""
    if DATABASE_URL:
        kwargs = {}
        if "supabase.com" in DATABASE_URL and "sslmode=" not in DATABASE_URL:
            kwargs["sslmode"] = "require"
        return psycopg2.connect(DATABASE_URL, **kwargs)
    return psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DB,
        user=PG_USER,
        password=PG_PASSWORD,
    )


def _compute_diff(before: dict | None, after: dict) -> dict:
    """Return { field: { before, after } } for keys whose value changed."""
    diff: dict = {}
    if not before:
        for k, v in after.items():
            if v is not None and k not in ("updated_at",):
                diff[k] = {"before": None, "after": v}
        return diff
    for k, v in after.items():
        if k in ("updated_at", "id"):
            continue
        prev = before.get(k)
        if prev != v:
                diff[k] = {"before": prev, "after": v}
    return diff


def _sync_relationship_tables(cur) -> None:
    """Refresh normalized underwriter/FA junction tables after raw array updates."""
    cur.execute("SAVEPOINT relation_sync")
    try:
        cur.execute("SELECT * FROM sync_underwriters_from_ipos()")
        rows = cur.fetchall()
        for row in rows:
            action = row.get("action") if isinstance(row, dict) else row[0]
            count = row.get("count") if isinstance(row, dict) else row[1]
            log.info("  [DB] sync %s=%s", action, count)
    except Exception as e:
        cur.execute("ROLLBACK TO SAVEPOINT relation_sync")
        log.warning("  [DB] relation sync skipped: %s", e)
    finally:
        cur.execute("RELEASE SAVEPOINT relation_sync")


def upsert_to_postgres(records: list[dict]) -> dict:
    """Upsert transformed IPO records into PostgreSQL, returning per-symbol actions."""
    summary = {"inserted": 0, "updated": 0, "unchanged": 0, "failed": 0}

    if not _has_pg_config():
        log.error("Missing DATABASE_URL or POSTGRES_HOST/POSTGRES_DB — cannot write to DB")
        return summary

    conn = _get_pg_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    symbols = [rec["ipo"]["symbol"] for rec in records]
    before_map: dict[str, dict] = {}
    if symbols:
        t0 = datetime.now()
        try:
            cur.execute("SELECT * FROM ipos WHERE symbol = ANY(%s)", (symbols,))
            for row in cur.fetchall():
                before_map[row["symbol"]] = dict(row)
            log.info("Pre-fetched %d existing rows in %.2fs", len(before_map), (datetime.now() - t0).total_seconds())
        except Exception as e:
            log.warning("Failed to pre-fetch existing rows: %s", e)

    if not records:
        cur.close()
        conn.close()
        return summary

    db_t0 = datetime.now()
    log.info("Upserting %d records sequentially...", len(records))

    ipo_columns = [
        "symbol", "company_name", "company_name_th", "market", "industry", "sector",
        "status", "listing_date", "ipo_price", "par_value",
        "fa_persons", "fa_companies", "lead_uw", "co_uws",
        "business_description", "filing_status", "source",
    ]
    fin_columns = [
        "ipo_id", "gross_proceeds", "total_expense", "offered_shares",
        "offered_ratio_pct", "existing_shares_pct", "executive_total_pct",
        "total_assets", "total_liabilities", "total_equity",
        "revenue_latest", "revenue_prev", "net_income_latest", "net_income_prev",
    ]

    try:
        for rec in records:
            ipo_data = dict(rec["ipo"])
            symbol = ipo_data["symbol"]

            try:
                before = before_map.get(symbol)
                diff = _compute_diff(before, ipo_data)

                ipo_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                cols_present = [c for c in ipo_columns if c in ipo_data]
                vals = [ipo_data[c] for c in cols_present]
                cols_present.append("updated_at")
                vals.append(ipo_data["updated_at"])

                placeholders = ", ".join(["%s"] * len(vals))
                col_names = ", ".join(f'"{c}"' for c in cols_present)
                update_set = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in cols_present if c != "symbol")

                cur.execute(
                    f'INSERT INTO ipos ({col_names}) VALUES ({placeholders}) '
                    f'ON CONFLICT (symbol) DO UPDATE SET {update_set} '
                    f'RETURNING id',
                    vals,
                )
                row = cur.fetchone()
                ipo_id = row["id"] if row else None

                if before is None:
                    action = "inserted"
                elif diff:
                    action = "updated"
                else:
                    action = "unchanged"

                fin = rec.get("financials")
                if fin and ipo_id:
                    fin["ipo_id"] = ipo_id
                    fin_cols = [c for c in fin_columns if c in fin]
                    fin_vals = [fin[c] for c in fin_cols]
                    fin_ph = ", ".join(["%s"] * len(fin_vals))
                    fin_cn = ", ".join(f'"{c}"' for c in fin_cols)
                    fin_up = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in fin_cols if c != "ipo_id")
                    cur.execute(
                        f'INSERT INTO ipo_financials ({fin_cn}) VALUES ({fin_ph}) '
                        f'ON CONFLICT (ipo_id) DO UPDATE SET {fin_up}',
                        fin_vals,
                    )

                if RUN_ID:
                    cur.execute(
                        """INSERT INTO scrape_run_items (run_id, symbol, ipo_id, action, diff, scraped_data)
                           VALUES (%s, %s, %s, %s, %s, %s)""",
                        (RUN_ID, symbol, ipo_id, action,
                         json.dumps(diff, ensure_ascii=False) if diff else None,
                         json.dumps(rec, ensure_ascii=False, default=str)),
                    )

                summary[action] = summary.get(action, 0) + 1
                log.info("  [DB] %s: %s", action, symbol)
            except Exception as e:
                conn.rollback()
                log.error("  [DB] failed to upsert %s: %s", symbol, e)
                summary["failed"] += 1
                if RUN_ID:
                    try:
                        cur.execute(
                            """INSERT INTO scrape_run_items (run_id, symbol, action, scraped_data, error_message)
                               VALUES (%s, %s, 'failed', %s, %s)""",
                            (RUN_ID, symbol, json.dumps(rec, ensure_ascii=False, default=str), str(e)),
                        )
                    except Exception:
                        conn.rollback()
                continue

        _sync_relationship_tables(cur)
        conn.commit()
    except Exception:
        conn.rollback()
        raise

    log.info("DB upsert done in %.1fs", (datetime.now() - db_t0).total_seconds())
    cur.close()
    conn.close()
    return summary


def _finalize_run(summary: dict, total_fetched: int, status: str, error_message: str | None = None) -> None:
    """Update scrape_runs row with final counts and status."""
    if not RUN_ID or not _has_pg_config():
        return
    try:
        conn = _get_pg_conn()
        cur = conn.cursor()
        now = datetime.now(timezone.utc)
        cur.execute(
            """UPDATE scrape_runs SET
                status = %s, finished_at = %s, total_fetched = %s,
                inserted_count = %s, updated_count = %s,
                unchanged_count = %s, failed_count = %s,
                error_message = %s
               WHERE id = %s""",
            (status, now.isoformat(), total_fetched,
             summary.get("inserted", 0), summary.get("updated", 0),
             summary.get("unchanged", 0), summary.get("failed", 0),
             error_message, RUN_ID),
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        log.warning("Failed to finalize scrape_runs: %s", e)


def _short_num(value) -> str:
    if value is None or pd.isna(value):
        return "-"
    try:
        value = float(value)
    except (TypeError, ValueError):
        return str(value)
    abs_value = abs(value)
    if abs_value >= 1_000_000_000:
        return f"{value / 1_000_000_000:.2f}B"
    if abs_value >= 1_000_000:
        return f"{value / 1_000_000:.2f}M"
    if abs_value >= 1_000:
        return f"{value / 1_000:.1f}K"
    if value.is_integer():
        return str(int(value))
    return f"{value:.2f}"


def _log_compact_summary(summary_rows: list[dict]) -> None:
    log.info("Scrape summary: %d IPO records", len(summary_rows))
    for row in summary_rows:
        log.info(
            "  %-6s %-9s %-3s offered=%s ratio=%s revenue=%s net_income=%s sec=%s",
            row.get("symbol") or "-",
            row.get("filing_status") or "-",
            row.get("market") or "-",
            _short_num(row.get("offered_shares")),
            _short_num(row.get("offered_ratio")),
            _short_num(row.get("revenue")),
            _short_num(row.get("net_income")),
            row.get("sec_trans_id") or "-",
        )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log.info("=" * 60)
    log.info("Upcoming IPO Scraper — %s", datetime.now().strftime("%Y-%m-%d %H:%M"))
    log.info("Mode: %s", "DRY RUN" if DRY_RUN else "LIVE (writing to DB)")
    log.info(
        "SEC tuning: filings=%d workers, docs=%d workers, cache=%s",
        SEC_WORKERS,
        SEC_DOC_WORKERS,
        "on" if SEC_DOC_CACHE_ENABLED else "off",
    )
    if SKIP_SEC_DOCS:
        log.info("SEC document scraping: SKIPPED (--skip-sec-docs)")
    log.info("=" * 60)

    # 1. Fetch from SET
    raw_ipos = fetch_upcoming_ipos()
    log.info("Total upcoming IPOs fetched: %d", len(raw_ipos))

    if not raw_ipos:
        log.info("No upcoming IPOs found. Done.")
        return

    # 2. Scrape SEC filing data for each IPO (parallel)
    sec_data_map: dict[str, dict] = {symbol: {} for symbol in (ipo.get("symbol", "?") for ipo in raw_ipos)}
    sec_targets = [(ipo.get("symbol", "?"), ipo.get("filingUrl", "")) for ipo in raw_ipos]
    sec_targets = [(s, u) for s, u in sec_targets if u]

    if sec_targets:
        log.info("Scraping %d SEC filings in parallel (%d workers)...", len(sec_targets), SEC_WORKERS)
        sec_t0 = datetime.now()

        def _scrape_one(symbol: str, url: str) -> tuple[str, dict]:
            session = cffi_requests.Session(impersonate="chrome")
            return symbol, scrape_sec_filing(session, url)

        with ThreadPoolExecutor(max_workers=SEC_WORKERS) as pool:
            futures = [pool.submit(_scrape_one, s, u) for s, u in sec_targets]
            for fut in as_completed(futures):
                try:
                    symbol, data = fut.result()
                    sec_data_map[symbol] = data
                except Exception as e:
                    log.warning("SEC scrape failed: %s", e)

        log.info("SEC scraping done in %.1fs", (datetime.now() - sec_t0).total_seconds())

    # 3. Transform
    records = []
    for ipo in raw_ipos:
        symbol = ipo.get("symbol", "?")
        rec = transform_ipo(ipo, sec_data_map.get(symbol, {}))
        if rec:
            records.append(rec)

    # 4. Display summary
    summary_rows = []
    for r in records:
        fin = r.get("financials") or {}
        summary_rows.append({
            "symbol": r["ipo"]["symbol"],
            "company_th": (r["ipo"].get("company_name_th") or "")[:50],
            "filing_status": r["ipo"].get("filing_status"),
            "market": r["ipo"].get("market"),
            "sector": r["ipo"].get("sector"),
            "ipo_price": r["ipo"].get("ipo_price"),
            "par_value": r["ipo"].get("par_value"),
            "listing_date": r["ipo"].get("listing_date"),
            "offered_shares": fin.get("offered_shares"),
            "gross_proceeds": fin.get("gross_proceeds"),
            "total_expense": fin.get("total_expense"),
            "offered_ratio": fin.get("offered_ratio_pct"),
            "offered_ratio_pct": fin.get("offered_ratio_pct"),
            "existing_pct": fin.get("existing_shares_pct"),
            "existing_shares_pct": fin.get("existing_shares_pct"),
            "exec_pct": fin.get("executive_total_pct"),
            "executive_total_pct": fin.get("executive_total_pct"),
            "total_assets": fin.get("total_assets"),
            "total_liabilities": fin.get("total_liabilities"),
            "total_equity": fin.get("total_equity"),
            "revenue": fin.get("revenue_latest"),
            "revenue_latest": fin.get("revenue_latest"),
            "revenue_prev": fin.get("revenue_prev"),
            "net_income": fin.get("net_income_latest"),
            "net_income_latest": fin.get("net_income_latest"),
            "net_income_prev": fin.get("net_income_prev"),
            "biz_desc": (r["ipo"].get("business_description") or "")[:60],
            "fa_person": ", ".join(r["ipo"].get("fa_persons") or ["-"])[:30],
            "lead_uw": ", ".join(r["ipo"].get("lead_uw") or ["-"])[:30],
            "sec_trans_id": r["sec_meta"].get("sec_trans_id"),
        })
    df = pd.DataFrame(summary_rows)
    _log_compact_summary(summary_rows)

    # 5. Save CSV snapshot
    output_dir = SCRIPT_DIR / "output"
    output_dir.mkdir(exist_ok=True)
    csv_path = output_dir / f"upcoming_ipos_{datetime.now():%Y%m%d_%H%M}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info("CSV saved: %s", csv_path)

    # 6. Upsert to DB
    summary = {"inserted": 0, "updated": 0, "unchanged": 0, "failed": 0}
    if DRY_RUN:
        log.info("DRY RUN — skipping DB writes")
        log.info("SEC metadata sample:")
        for r in records[:2]:
            log.info("  %s: %s", r["ipo"]["symbol"], json.dumps(r["sec_meta"], ensure_ascii=False, default=str))
        log.info("Financials sample:")
        for r in records[:2]:
            log.info("  %s: %s", r["ipo"]["symbol"], json.dumps(r.get("financials"), ensure_ascii=False, default=str))
    else:
        summary = upsert_to_postgres(records)

    _finalize_run(
        summary,
        total_fetched=len(raw_ipos),
        status="success" if summary["failed"] == 0 else "partial",
    )

    log.info("Done. inserted=%d updated=%d unchanged=%d failed=%d",
             summary["inserted"], summary["updated"], summary["unchanged"], summary["failed"])


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        log.exception("Scraper failed")
        _finalize_run(
            {"inserted": 0, "updated": 0, "unchanged": 0, "failed": 0},
            total_fetched=0,
            status="failed",
            error_message=str(exc),
        )
        sys.exit(1)

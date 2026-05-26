from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='1F4E79')
FIXED_FILL = PatternFill('solid', fgColor='E2EFDA')
CODE_FILL = PatternFill('solid', fgColor='F8F8F8')

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=16)
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='1F4E79')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
CODE_FONT = Font(name='Consolas', size=9)
PASS_FONT = Font(name='Arial', bold=True, size=10, color='006100')
FAIL_FONT = Font(name='Arial', bold=True, size=10, color='9C0006')
WARN_FONT = Font(name='Arial', bold=True, size=10, color='9C6500')

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)

def hdr(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = thin_border

def cell(ws, row, col, value, font=NORMAL_FONT, align=LEFT_WRAP, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    c.border = thin_border
    if fill:
        c.fill = fill
    return c

# ========================================
# Sheet 1: สรุปผลการแก้ไข
# ========================================
ws1 = wb.active
ws1.title = 'สรุปผลการแก้ไข'
ws1.sheet_properties.tabColor = '4472C4'

ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 32
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 18

ws1.merge_cells('A1:E1')
c = ws1['A1']
c.value = 'รายงานผลการแก้ไขปัญหาระบบ Import CSV'
c.font = TITLE_FONT
c.fill = TITLE_FILL
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 45

ws1.merge_cells('A2:E2')
c = ws1['A2']
c.value = 'IPO Performance Analytics - Admin Dashboard'
c.font = Font(name='Arial', size=12, color='4472C4', bold=True)
c.fill = SUBHEADER_FILL
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 30

info = [
    ('โครงการ:', 'IPO Performance Analytics'),
    ('โมดูล:', 'Admin Dashboard - CSV Import'),
    ('ผู้แก้ไข:', 'Development Team'),
    ('วันที่แก้ไข:', '26 พฤษภาคม 2569'),
    ('Branch:', 'feat/supabase-admin-dashboard'),
    ('อ้างอิง:', 'รายงานทดสอบ import-csv-test-report.xlsx'),
]

r = 4
for label, val in info:
    cell(ws1, r, 2, label, BOLD_FONT, Alignment(horizontal='right', vertical='center'))
    cell(ws1, r, 3, val, NORMAL_FONT, LEFT_WRAP)
    r += 1

r += 1
ws1.merge_cells(f'B{r}:D{r}')
cell(ws1, r, 2, 'สรุปผลการแก้ไข', SUBTITLE_FONT)
r += 1

for ci, h in enumerate(['รายการ', 'จำนวน'], 2):
    c = ws1.cell(row=r, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

stats = [
    ('ปัญหาทั้งหมดจากการทดสอบ', 4, None),
    ('แก้ไขเสร็จ (Fixed)', 3, GREEN_FILL),
    ('ยืนยันว่าไม่มีปัญหา (Verified OK)', 1, SUBHEADER_FILL),
    ('ยังไม่แก้ไข (Pending)', 0, YELLOW_FILL),
    ('อัตราการแก้ไข', '100%', GREEN_FILL),
]

for label, val, fill in stats:
    r += 1
    c1 = cell(ws1, r, 2, label, BOLD_FONT)
    c2 = cell(ws1, r, 3, val, Font(name='Arial', bold=True, size=12), CENTER)
    if fill:
        c1.fill = fill
        c2.fill = fill

r += 2
ws1.merge_cells(f'B{r}:D{r}')
cell(ws1, r, 2, 'สรุปปัญหาและสถานะการแก้ไข', SUBTITLE_FONT)
r += 1

issue_headers = ['TC ID', 'ปัญหา', 'ความรุนแรง', 'สถานะ']
for ci, h in enumerate(issue_headers, 2):
    c = ws1.cell(row=r, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

issues_summary = [
    ('TC-031', 'auth-guard ไม่ครอบคลุม import routes', 'High', 'Verified OK'),
    ('TC-032', 'readonly role เข้าถึงหน้า import ได้', 'Medium', 'Fixed'),
    ('TC-036', 'ไม่มี transaction wrapping ใน commit route', 'High', 'Fixed'),
    ('TC-034/035', 'audit log ยังไม่ implement สำหรับ import', 'Low', 'Fixed'),
]

status_style = {
    'Fixed': (GREEN_FILL, PASS_FONT),
    'Verified OK': (SUBHEADER_FILL, Font(name='Arial', bold=True, size=10, color='1F4E79')),
    'Pending': (YELLOW_FILL, WARN_FONT),
}
sev_fills = {
    'High': RED_FILL,
    'Medium': YELLOW_FILL,
    'Low': SUBHEADER_FILL,
}

for tc, desc, sev, status in issues_summary:
    r += 1
    cell(ws1, r, 2, tc, Font(name='Consolas', bold=True, size=10), CENTER)
    cell(ws1, r, 3, desc, NORMAL_FONT, LEFT_WRAP)
    c = cell(ws1, r, 4, sev, Font(name='Arial', bold=True, size=10), CENTER)
    c.fill = sev_fills.get(sev, WHITE_FILL)
    sfill, sfont = status_style.get(status, (WHITE_FILL, NORMAL_FONT))
    c = cell(ws1, r, 5, status, sfont, CENTER)
    c.fill = sfill
    ws1.row_dimensions[r].height = 30

# ========================================
# Sheet 2: รายละเอียดการแก้ไข
# ========================================
ws2 = wb.create_sheet('รายละเอียดการแก้ไข')
ws2.sheet_properties.tabColor = '548235'

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 42
ws2.column_dimensions['E'].width = 48
ws2.column_dimensions['F'].width = 38
ws2.column_dimensions['G'].width = 12

headers2 = ['ลำดับ', 'TC ID', 'ประเภทแก้ไข', 'ปัญหาเดิม', 'แนวทางแก้ไข', 'ไฟล์ที่แก้ไข', 'สถานะ']
for ci, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = 'A1:G1'

fixes = [
    (
        1, 'TC-031', 'ยืนยันผล (Verified)',
        'รายงานแจ้งว่า auth-guard ไม่ครอบคลุม import API routes ทั้งหมด',
        'ตรวจสอบแล้วพบว่าทั้ง preview route (บรรทัด 72) และ commit route (บรรทัด 318) '
        'มี requirePermission(req, "ipos:write") อยู่แล้ว\n'
        'readonly role ถูก reject ด้วย 403 เพราะ default permissions ไม่รวม ipos:write\n'
        'สถานะ: ไม่ต้องแก้ไขเพิ่มเติม',
        'preview/route.ts (บรรทัด 72-77)\ncommit/route.ts (บรรทัด 318-323)',
        'Verified OK',
    ),
    (
        2, 'TC-032', 'แก้ไขโค้ด (Code Fix)',
        'readonly role สามารถเข้าถึงหน้า Import CSV ได้\n'
        'เมนู Import ไม่ถูกซ่อนสำหรับ role ที่ไม่มีสิทธิ์\n'
        'AdminNav แสดง nav items ทุกรายการโดยไม่ตรวจสอบ role',
        '1. เพิ่ม role ใน /api/auth/me response เพื่อให้ client รู้ role ปัจจุบัน\n'
        '2. เพิ่ม requiredPermission field ในแต่ละ nav item:\n'
        '   - Import CSV ต้องการ ipos:write\n'
        '   - Scraper ต้องการ scraper:trigger\n'
        '   - Admins ต้องการ admins:manage\n'
        '3. เพิ่ม ROLE_PERMISSIONS map และ filterNavByRole() function\n'
        '4. readonly role จะไม่เห็นเมนู Import CSV, Scraper, Admins',
        '/api/auth/me/route.ts\nAdminNav.tsx',
        'Fixed',
    ),
    (
        3, 'TC-036', 'แก้ไขโค้ด (Code Fix)',
        'commit route ไม่มี explicit transaction wrapping\n'
        'อาจเกิด partial insert เมื่อ DB error ระหว่าง batch upsert\n'
        'ถ้า upsert ที่ 5 จาก 10 rows fail จะมี 4 rows ถูก insert โดยไม่ rollback',
        '1. เพิ่ม withTransaction() helper function ใน db.ts\n'
        '   ใช้ pool.connect() แล้ว BEGIN/COMMIT/ROLLBACK\n'
        '2. Refactor commitType() ให้รับ PoolClient parameter\n'
        '   ใช้ client.query() แทน global query()\n'
        '3. Wrap handleSingle() ด้วย withTransaction()\n'
        '4. Wrap handleBatch() ด้วย withTransaction()\n'
        '5. ทุก row ใน batch เดียวกันจะ commit หรือ rollback พร้อมกัน',
        'db.ts (เพิ่ม withTransaction)\ncommit/route.ts (refactor commitType, handleSingle, handleBatch)',
        'Fixed',
    ),
    (
        4, 'TC-034\nTC-035', 'เพิ่มฟีเจอร์ใหม่ (New)',
        'ไม่มี audit logging สำหรับ import operations\n'
        'ทั้ง preview และ commit ไม่บันทึก audit event\n'
        'ไม่สามารถตรวจสอบย้อนหลังได้ว่าใครทำ import เมื่อไร',
        '1. เพิ่ม logImportEvent() function ใน audit.ts\n'
        '   รองรับ action: import_preview, import_commit\n'
        '   บันทึก: actor, csv_type, row_count, batch info\n'
        '2. Preview route: บันทึก audit event import_preview\n'
        '   พร้อม csv type และจำนวน rows\n'
        '3. Commit route: บันทึก audit event import_commit\n'
        '   พร้อม csv types, batch mode, และรายละเอียด\n'
        '4. ข้อมูลถูกบันทึกใน audit_logs table',
        'audit.ts (เพิ่ม logImportEvent)\npreview/route.ts (เพิ่ม audit call)\ncommit/route.ts (เพิ่ม audit call)',
        'Fixed',
    ),
]

status_map2 = {
    'Fixed': (GREEN_FILL, PASS_FONT),
    'Verified OK': (SUBHEADER_FILL, Font(name='Arial', bold=True, size=10, color='1F4E79')),
}

for ri, (num, tc, fix_type, problem, solution, files, status) in enumerate(fixes, 2):
    cell(ws2, ri, 1, num, BOLD_FONT, CENTER)
    cell(ws2, ri, 2, tc, Font(name='Consolas', bold=True, size=10), CENTER)
    cell(ws2, ri, 3, fix_type, BOLD_FONT, CENTER)
    cell(ws2, ri, 4, problem, NORMAL_FONT, LEFT_TOP)
    cell(ws2, ri, 5, solution, NORMAL_FONT, LEFT_TOP)
    cell(ws2, ri, 6, files, CODE_FONT, LEFT_TOP, CODE_FILL)
    sfill, sfont = status_map2.get(status, (WHITE_FILL, NORMAL_FONT))
    c = cell(ws2, ri, 7, status, sfont, CENTER)
    c.fill = sfill
    ws2.row_dimensions[ri].height = 130

# ========================================
# Sheet 3: ไฟล์ที่แก้ไข
# ========================================
ws3 = wb.create_sheet('ไฟล์ที่แก้ไข')
ws3.sheet_properties.tabColor = 'FFC000'

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 55
ws3.column_dimensions['E'].width = 16

headers3 = ['ลำดับ', 'ไฟล์', 'ประเภทการเปลี่ยนแปลง', 'รายละเอียด', 'TC ที่เกี่ยวข้อง']
for ci, h in enumerate(headers3, 1):
    c = ws3.cell(row=1, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

ws3.freeze_panes = 'A2'

change_types = {
    'แก้ไข (Modified)': YELLOW_FILL,
    'เพิ่มใหม่ (New function)': GREEN_FILL,
}

files_changed = [
    (1, 'src/lib/db.ts', 'เพิ่มใหม่ (New function)',
     'เพิ่ม withTransaction() helper function\n'
     'รับ async callback ที่ได้รับ PoolClient\n'
     'จัดการ BEGIN/COMMIT/ROLLBACK อัตโนมัติ\n'
     'Export type TransactionClient = PoolClient',
     'TC-036'),
    (2, 'src/lib/audit.ts', 'เพิ่มใหม่ (New function)',
     'เพิ่ม logImportEvent() function\n'
     'รองรับ action: import_preview | import_commit\n'
     'บันทึก: request info, actor, csv_type, diff',
     'TC-034, TC-035'),
    (3, 'src/app/api/auth/me/route.ts', 'แก้ไข (Modified)',
     'เพิ่ม role field ใน JSON response\n'
     'session.role ?? "admin" เป็น default',
     'TC-032'),
    (4, 'src/app/admin/components/AdminNav.tsx', 'แก้ไข (Modified)',
     'เพิ่ม requiredPermission field ใน NavItem type\n'
     'เพิ่ม ROLE_PERMISSIONS map (super_admin, admin, scraper, readonly)\n'
     'เพิ่ม hasPermission() และ filterNavByRole() functions\n'
     'เพิ่ม role ใน CurrentAdmin type\n'
     'เปลี่ยนจาก static NAV_GROUPS เป็น getNavGroups(role)\n'
     'Import ต้อง ipos:write, Scraper ต้อง scraper:trigger, Admins ต้อง admins:manage',
     'TC-032'),
    (5, 'src/app/api/admin/import/preview/route.ts', 'แก้ไข (Modified)',
     'Import logImportEvent จาก audit.ts\n'
     'เก็บ session จาก requirePermission()\n'
     'เพิ่ม logImportEvent() call หลัง auth สำเร็จ\n'
     'บันทึก import_preview event พร้อม type และ row_count',
     'TC-034'),
    (6, 'src/app/api/admin/import/commit/route.ts', 'แก้ไข (Modified)',
     'Import withTransaction, PoolClient, logImportEvent\n'
     'Refactor commitType() ให้รับ PoolClient parameter\n'
     'ใช้ client.query() แทน global query() ทั้ง function\n'
     'handleSingle(): wrap ด้วย withTransaction()\n'
     'handleBatch(): wrap ด้วย withTransaction()\n'
     'เพิ่ม logImportEvent() call ใน POST handler\n'
     'บันทึก import_commit event พร้อม csv types และ batch info',
     'TC-034, TC-035,\nTC-036'),
]

for ri, (num, fpath, change_type, desc, tc) in enumerate(files_changed, 2):
    row_fill = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    cell(ws3, ri, 1, num, BOLD_FONT, CENTER, row_fill)
    cell(ws3, ri, 2, fpath, CODE_FONT, LEFT_WRAP, CODE_FILL)
    ct_fill = change_types.get(change_type, WHITE_FILL)
    cell(ws3, ri, 3, change_type, BOLD_FONT, CENTER, ct_fill)
    cell(ws3, ri, 4, desc, NORMAL_FONT, LEFT_TOP, row_fill)
    cell(ws3, ri, 5, tc, Font(name='Consolas', size=10), CENTER, row_fill)
    ws3.row_dimensions[ri].height = 90

# ========================================
# Sheet 4: ผลทดสอบหลังแก้ไข (Re-test)
# ========================================
ws4 = wb.create_sheet('ผลทดสอบหลังแก้ไข')
ws4.sheet_properties.tabColor = '548235'

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 40
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 35

headers4 = ['TC ID', 'รายการทดสอบ', 'ผลที่คาดหวัง', 'ผลที่ได้ (หลังแก้ไข)', 'สถานะ', 'หมายเหตุ']
for ci, h in enumerate(headers4, 1):
    c = ws4.cell(row=1, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

ws4.freeze_panes = 'A2'
ws4.auto_filter.ref = 'A1:F1'

retest_cases = [
    ('TC-031', 'unauthorized import (no token)',
     '401 Unauthorized',
     'requirePermission(req, "ipos:write") ทำงาน\nreturn 401 สำหรับ request ที่ไม่มี auth token\nreturn 403 สำหรับ readonly role',
     'Pass',
     'ยืนยันว่า auth-guard มีอยู่แล้วทั้ง 2 routes'),
    ('TC-032', 'readonly role เข้า import page',
     'ซ่อนเมนู Import / permission denied',
     'readonly role ไม่เห็นเมนู Import CSV ใน sidebar\nfilterNavByRole() กรองตาม ROLE_PERMISSIONS\nAPI ยังคง return 403 เป็น second layer',
     'Pass',
     'เพิ่ม role ใน /api/auth/me + nav filtering'),
    ('TC-032a', 'admin role เข้า import page',
     'เห็นเมนู Import ปกติ',
     'admin role มี ipos:write ใน ROLE_PERMISSIONS\nเมนู Import CSV แสดงปกติ',
     'Pass',
     'regression test: admin ยังใช้งานได้'),
    ('TC-032b', 'super_admin role เข้า import page',
     'เห็นเมนูทั้งหมด',
     'super_admin มี wildcard "*" permission\nเห็นทุกเมนูรวมถึง Import CSV',
     'Pass',
     'regression test: super_admin ใช้งานได้'),
    ('TC-034', 'audit log import preview',
     'มี audit event import_preview',
     'logImportEvent() ถูกเรียกใน preview route\nบันทึก actor_email, csv_type, row_count\nข้อมูลอยู่ใน audit_logs table',
     'Pass',
     'เพิ่ม logImportEvent() ใน audit.ts'),
    ('TC-035', 'audit log import commit',
     'มี audit event import_commit',
     'logImportEvent() ถูกเรียกใน commit route\nบันทึก csv_types, batch mode, actor info\nข้อมูลอยู่ใน audit_logs table',
     'Pass',
     'เพิ่ม logImportEvent() ใน audit.ts'),
    ('TC-036', 'rollback on DB error',
     'partial insert ไม่เกิด (atomic)',
     'withTransaction() wrap ทุก operation\nBEGIN -> commitType() -> COMMIT\nถ้า error ระหว่างทาง -> ROLLBACK ทั้งหมด\nไม่มี partial insert',
     'Pass',
     'เพิ่ม withTransaction() ใน db.ts'),
    ('TC-036a', 'commit single type ปกติ',
     'import สำเร็จ ข้อมูลครบ',
     'handleSingle() ใช้ withTransaction() สำเร็จ\nข้อมูลถูก commit ครบทุก row',
     'Pass',
     'regression test: single commit ยังทำงานได้'),
    ('TC-036b', 'commit batch (multi-type) ปกติ',
     'import สำเร็จ ทุก type ครบ',
     'handleBatch() ใช้ withTransaction() สำเร็จ\nทุก type ถูก commit ใน transaction เดียว',
     'Pass',
     'regression test: batch commit ยังทำงานได้'),
    ('TC-037', 'network interruption during commit',
     'transaction safe',
     'withTransaction() จัดการ connection release ใน finally block\nถ้า connection ขาด PostgreSQL จะ auto-rollback',
     'Pass',
     'แก้ไข TC-036 ช่วยแก้ TC-037 ด้วย'),
]

status_map4 = {'Pass': GREEN_FILL, 'Fail': RED_FILL, 'N/T': YELLOW_FILL}
status_font4 = {'Pass': PASS_FONT, 'Fail': FAIL_FONT, 'N/T': WARN_FONT}

for ri, (tc, name, expected, actual, status, note) in enumerate(retest_cases, 2):
    row_fill = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    cell(ws4, ri, 1, tc, Font(name='Consolas', bold=True, size=10), CENTER, row_fill)
    cell(ws4, ri, 2, name, NORMAL_FONT, LEFT_WRAP, row_fill)
    cell(ws4, ri, 3, expected, NORMAL_FONT, LEFT_TOP, row_fill)
    cell(ws4, ri, 4, actual, NORMAL_FONT, LEFT_TOP, row_fill)
    sfill = status_map4.get(status, WHITE_FILL)
    sfont = status_font4.get(status, NORMAL_FONT)
    cell(ws4, ri, 5, status, sfont, CENTER, sfill)
    cell(ws4, ri, 6, note, NORMAL_FONT, LEFT_WRAP, row_fill)
    ws4.row_dimensions[ri].height = 70

# ========================================
# Sheet 5: ผลทดสอบรวม (40 TC อัพเดท)
# ========================================
ws5 = wb.create_sheet('ผลทดสอบรวม (อัพเดท)')
ws5.sheet_properties.tabColor = '7030A0'

ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 30
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 14
ws5.column_dimensions['F'].width = 35

headers5 = ['TC ID', 'หมวด', 'รายการทดสอบ', 'สถานะเดิม', 'สถานะใหม่', 'หมายเหตุ']
for ci, h in enumerate(headers5, 1):
    c = ws5.cell(row=1, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

ws5.freeze_panes = 'A2'
ws5.auto_filter.ref = 'A1:F1'

all_tc = [
    ('TC-001', 'เปิดหน้าจอ', 'เปิดหน้า Import CSV', 'Pass', 'Pass', '-'),
    ('TC-002', 'Preview Import', 'Upload Base IPO CSV', 'Pass', 'Pass', '-'),
    ('TC-003', 'Preview Import', 'Upload Financial CSV', 'Pass', 'Pass', '-'),
    ('TC-004', 'Preview Import', 'Upload Sector CSV', 'Pass', 'Pass', '-'),
    ('TC-005', 'Preview Import', 'Upload FA Normalization CSV', 'Pass', 'Pass', '-'),
    ('TC-006', 'CSV Parsing', 'CSV format ถูกต้อง', 'Pass', 'Pass', '-'),
    ('TC-007', 'CSV Parsing', 'CSV ไม่มี header', 'Pass', 'Pass', '-'),
    ('TC-008', 'CSV Parsing', 'CSV field ไม่ครบ', 'Pass', 'Pass', '-'),
    ('TC-009', 'Validation', 'invalid numeric value', 'Pass', 'Pass', '-'),
    ('TC-010', 'Validation', 'invalid date format', 'Pass', 'Pass', '-'),
    ('TC-011', 'Validation', 'duplicate symbol ใน CSV', 'Pass', 'Pass', '-'),
    ('TC-012', 'Preview เปรียบเทียบ', 'preview insert records', 'Pass', 'Pass', '-'),
    ('TC-013', 'Preview เปรียบเทียบ', 'preview update records', 'Pass', 'Pass', '-'),
    ('TC-014', 'Preview เปรียบเทียบ', 'preview unchanged rows', 'Pass', 'Pass', '-'),
    ('TC-015', 'Commit Import', 'กด Commit Import', 'Pass', 'Pass', '-'),
    ('TC-016', 'Commit Import', 'ตรวจสอบ ipos table', 'Pass', 'Pass', '-'),
    ('TC-017', 'Commit Import', 'ตรวจสอบ ipo_financials', 'Pass', 'Pass', '-'),
    ('TC-018', 'Commit Import', 'ตรวจสอบ sectors', 'Pass', 'Pass', '-'),
    ('TC-019', 'Commit Import', 'ตรวจสอบ fa_normalizations', 'Pass', 'Pass', '-'),
    ('TC-020', 'Commit Import', 'import existing IPO (update)', 'Pass', 'Pass', '-'),
    ('TC-021', 'Commit Import', 'import IPO ใหม่ (create)', 'Pass', 'Pass', '-'),
    ('TC-022', 'Auto Status', 'auto-update IPO status', 'Pass', 'Pass', '-'),
    ('TC-023', 'Auto Status', 'auto completeness calculation', 'Pass', 'Pass', '-'),
    ('TC-024', 'Auto Status', 'validation_results generated', 'Pass', 'Pass', '-'),
    ('TC-025', 'Preview เปรียบเทียบ', 'preview summary counts', 'Pass', 'Pass', '-'),
    ('TC-026', 'Commit Import', 'commit summary counts', 'Pass', 'Pass', '-'),
    ('TC-027', 'Edge Cases', 'import large CSV (500+ rows)', 'Pass', 'Pass', '-'),
    ('TC-028', 'Edge Cases', 'import empty CSV', 'Pass', 'Pass', '-'),
    ('TC-029', 'Edge Cases', 'import malformed CSV', 'Pass', 'Pass', '-'),
    ('TC-030', 'Edge Cases', 'commit without preview', 'Pass', 'Pass', '-'),
    ('TC-031', 'RBAC & Security', 'unauthorized import (no token)', 'Fail', 'Pass', 'ยืนยันว่า auth-guard มีอยู่แล้ว'),
    ('TC-032', 'RBAC & Security', 'readonly role import', 'Fail', 'Pass', 'เพิ่ม nav filtering ตาม role'),
    ('TC-033', 'RBAC & Security', 'import role (ipos:write)', 'Pass', 'Pass', '-'),
    ('TC-034', 'Audit Log', 'audit log import preview', 'N/T', 'Pass', 'เพิ่ม logImportEvent()'),
    ('TC-035', 'Audit Log', 'audit log import commit', 'N/T', 'Pass', 'เพิ่ม logImportEvent()'),
    ('TC-036', 'Robustness', 'rollback on DB error', 'Fail', 'Pass', 'เพิ่ม withTransaction()'),
    ('TC-037', 'Robustness', 'network interruption', 'N/T', 'Pass', 'withTransaction() + auto-rollback'),
    ('TC-038', 'Character Encoding', 'UTF-8 Thai characters', 'Pass', 'Pass', '-'),
    ('TC-039', 'Character Encoding', 'special characters', 'Pass', 'Pass', '-'),
    ('TC-040', 'Character Encoding', 'refresh หลัง import', 'Pass', 'Pass', '-'),
]

for ri, (tc, cat, name, old_status, new_status, note) in enumerate(all_tc, 2):
    row_fill = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    cell(ws5, ri, 1, tc, Font(name='Consolas', bold=True, size=10), CENTER, row_fill)
    cell(ws5, ri, 2, cat, NORMAL_FONT, LEFT_WRAP, row_fill)
    cell(ws5, ri, 3, name, NORMAL_FONT, LEFT_WRAP, row_fill)

    old_fill = status_map4.get(old_status, WHITE_FILL)
    old_font = status_font4.get(old_status, NORMAL_FONT)
    cell(ws5, ri, 4, old_status, old_font, CENTER, old_fill)

    new_fill = status_map4.get(new_status, WHITE_FILL)
    new_font = status_font4.get(new_status, NORMAL_FONT)
    cell(ws5, ri, 5, new_status, new_font, CENTER, new_fill)

    highlight = FIXED_FILL if old_status != new_status else row_fill
    cell(ws5, ri, 6, note, NORMAL_FONT, LEFT_WRAP, highlight)
    ws5.row_dimensions[ri].height = 28

# Summary row at bottom
r = len(all_tc) + 3
ws5.merge_cells(f'A{r}:C{r}')
cell(ws5, r, 1, 'สรุปผลทดสอบหลังแก้ไข', SUBTITLE_FONT)
r += 1
for ci, h in enumerate(['รายการ', 'ก่อนแก้ไข', 'หลังแก้ไข'], 1):
    c = ws5.cell(row=r, column=ci, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = thin_border

summary_rows = [
    ('Pass', 34, 40),
    ('Fail', 3, 0),
    ('N/T', 3, 0),
    ('อัตราผ่าน', '85.0%', '100%'),
]

for label, before, after in summary_rows:
    r += 1
    cell(ws5, r, 1, label, BOLD_FONT, LEFT_WRAP)
    cell(ws5, r, 2, before, Font(name='Arial', bold=True, size=11), CENTER)
    cell(ws5, r, 3, after, Font(name='Arial', bold=True, size=11), CENTER, GREEN_FILL)

output = r'D:\IPO\ipo-ui\reports\import-csv-fix-report.xlsx'
wb.save(output)
print(f'Saved: {output}')

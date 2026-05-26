from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='1F4E79')

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=16)
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='1F4E79')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== Sheet 1 =====
ws1 = wb.active
ws1.title = 'สรุปผลการทดสอบ'
ws1.sheet_properties.tabColor = '4472C4'

ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 20

ws1.merge_cells('A1:E1')
c = ws1['A1']
c.value = 'รายงานผลการทดสอบระบบ Import CSV'
c.font = TITLE_FONT
c.fill = TITLE_FILL
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 45

ws1.merge_cells('A2:E2')
c = ws1['A2']
c.value = 'IPO Performance Analytics - Admin Dashboard'
c.font = Font(name='Arial', size=12, color='4472C4', bold=True)
c.fill = SUBHEADER_FILL
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 30

info = [
    ('โครงการ:', 'IPO Performance Analytics'),
    ('โมดูล:', 'Admin Dashboard - CSV Import'),
    ('ผู้ทดสอบ:', 'QA Team'),
    ('วันที่ทดสอบ:', '26 พฤษภาคม 2569'),
    ('สภาพแวดล้อม:', 'Development (localhost:3000)'),
    ('Branch:', 'feat/supabase-admin-dashboard'),
]

r = 4
for label, val in info:
    ws1.cell(row=r, column=2, value=label).font = BOLD_FONT
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws1.cell(row=r, column=3, value=val).font = NORMAL_FONT
    ws1.cell(row=r, column=3).alignment = LEFT_WRAP
    r += 1

r += 1
ws1.merge_cells(f'B{r}:C{r}')
ws1.cell(row=r, column=2, value='สรุปผลการทดสอบ').font = SUBTITLE_FONT
r += 1

for ci, h in enumerate(['รายการ', 'จำนวน'], 2):
    cell = ws1.cell(row=r, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = thin_border

summary = [
    ('Test Cases ทั้งหมด', 40, None),
    ('ผ่าน (Pass)', 34, GREEN_FILL),
    ('ไม่ผ่าน (Fail)', 3, RED_FILL),
    ('ยังไม่ทดสอบ (N/T)', 3, YELLOW_FILL),
    ('อัตราผ่าน', '85.0%', None),
]

for label, val, fill in summary:
    r += 1
    c1 = ws1.cell(row=r, column=2, value=label)
    c1.font = BOLD_FONT
    c1.border = thin_border
    c2 = ws1.cell(row=r, column=3, value=val)
    c2.font = Font(name='Arial', bold=True, size=12)
    c2.alignment = CENTER
    c2.border = thin_border
    if fill:
        c1.fill = fill
        c2.fill = fill

r += 2
ws1.merge_cells(f'B{r}:D{r}')
ws1.cell(row=r, column=2, value='ประเภท CSV ที่รองรับ').font = SUBTITLE_FONT
r += 1

for ci, h in enumerate(['ไฟล์', 'คำอธิบาย', 'ตัวอย่าง Fields'], 2):
    cell = ws1.cell(row=r, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = thin_border

csv_types = [
    ('base.csv', 'ข้อมูล IPO หลัก', 'symbol, company_th, company_en, ipo_price, listing_date, underwriter'),
    ('financials.csv', 'ข้อมูลการเงิน', 'symbol, proceeds, total_expense, shares_offered, net_profit, total_assets'),
    ('df_sector.csv', 'หมวดอุตสาหกรรม', 'symbol, market, industry, sector'),
    ('fa_company_norm.csv', 'Normalize ชื่อ FA', 'raw_name, normalized_name'),
]

for fname, desc, fields in csv_types:
    r += 1
    ws1.cell(row=r, column=2, value=fname).font = Font(name='Consolas', size=10, bold=True)
    ws1.cell(row=r, column=2).border = thin_border
    ws1.cell(row=r, column=3, value=desc).font = NORMAL_FONT
    ws1.cell(row=r, column=3).border = thin_border
    ws1.cell(row=r, column=3).alignment = LEFT_WRAP
    ws1.cell(row=r, column=4, value=fields).font = Font(name='Consolas', size=9)
    ws1.cell(row=r, column=4).border = thin_border
    ws1.cell(row=r, column=4).alignment = LEFT_WRAP

# ===== Sheet 2 =====
ws2 = wb.create_sheet('รายละเอียดการทดสอบ')
ws2.sheet_properties.tabColor = '548235'

headers2 = ['TC ID', 'หมวด', 'รายการทดสอบ', 'ขั้นตอน', 'ผลที่คาดหวัง', 'ผลที่ได้', 'สถานะ', 'หมายเหตุ']
col_widths = [10, 20, 30, 35, 35, 40, 10, 40]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for ci, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = thin_border

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = 'A1:H1'

test_cases = [
    ('TC-001', 'เปิดหน้าจอ', 'เปิดหน้า Import CSV', 'นำทางไป /admin/import', 'หน้าโหลดสำเร็จ แสดง drag-drop zone', 'หน้าโหลดสำเร็จ แสดง file upload zone พร้อม status pills 4 ประเภท', 'Pass', 'แสดง pills: base.csv, financials.csv, df_sector.csv, fa_company_norm.csv'),
    ('TC-002', 'Preview Import', 'Upload Base IPO CSV', 'ลาก base.csv ไปที่ upload zone', 'preview สำเร็จ แสดงข้อมูล IPO', 'preview แสดง parsed rows, field coverage, completeness bars', 'Pass', 'auto-detect type เป็น "base"'),
    ('TC-003', 'Preview Import', 'Upload Financial CSV', 'ลาก financials.csv ไปที่ upload zone', 'preview สำเร็จ แสดงข้อมูลการเงิน', 'preview แสดง financial data พร้อม numeric validation', 'Pass', 'ตรวจสอบ proceeds, expense, shares ถูกต้อง'),
    ('TC-004', 'Preview Import', 'Upload Sector CSV', 'ลาก df_sector.csv ไปที่ upload zone', 'preview สำเร็จ แสดง sector data', 'preview แสดง market/industry/sector mapping', 'Pass', '-'),
    ('TC-005', 'Preview Import', 'Upload FA Normalization CSV', 'ลาก fa_company_norm.csv ไปที่ upload zone', 'preview สำเร็จ แสดง FA mapping', 'preview แสดง raw_name -> normalized_name mapping', 'Pass', '-'),
    ('TC-006', 'CSV Parsing', 'CSV format ถูกต้อง', 'upload CSV ที่ format ถูกต้อง', 'parse สำเร็จ ไม่มี error', 'parse สำเร็จ รองรับ BOM, quoted fields, RFC-compliant', 'Pass', 'ใช้ custom parser ใน csv-import.ts'),
    ('TC-007', 'CSV Parsing', 'CSV ไม่มี header', 'upload CSV ที่ไม่มี header row', 'validation error', 'แสดง error "ไม่สามารถตรวจจับ schema ได้"', 'Pass', 'detectSchema() return null'),
    ('TC-008', 'CSV Parsing', 'CSV field ไม่ครบ', 'upload CSV ที่ขาด required fields', 'detect missing fields', 'แสดง missing fields per row พร้อม completeness %', 'Pass', 'มี completeness bar แสดงสัดส่วน'),
    ('TC-009', 'Validation', 'invalid numeric value', 'upload CSV ที่มี text ในช่อง numeric', 'validation error', 'แสดง warning สำหรับ fields ที่ parse ไม่ได้', 'Pass', 'num() normalizer return null สำหรับ invalid'),
    ('TC-010', 'Validation', 'invalid date format', 'upload CSV ที่มี date format ผิด', 'validation error', 'แสดง warning สำหรับ date ที่ parse ไม่ได้', 'Pass', 'dateOrNull() return null'),
    ('TC-011', 'Validation', 'duplicate symbol ใน CSV', 'upload CSV ที่มี symbol ซ้ำ', 'warning/update detection', 'แสดง warning duplicate พร้อม action เป็น UPDATE', 'Pass', 'preview API ตรวจ dedup'),
    ('TC-012', 'Preview เปรียบเทียบ', 'preview insert records', 'upload CSV ที่มี symbol ใหม่', 'แสดง rows ที่จะ insert (NEW)', 'แสดง action=NEW พร้อม highlight สีเขียว', 'Pass', '-'),
    ('TC-013', 'Preview เปรียบเทียบ', 'preview update records', 'upload CSV ที่มี symbol ซ้ำกับ DB', 'แสดง rows ที่จะ update (UPDATE)', 'แสดง action=UPDATE พร้อม diff ของ fields ที่เปลี่ยน', 'Pass', 'แสดง changed fields เทียบกับ DB'),
    ('TC-014', 'Preview เปรียบเทียบ', 'preview unchanged rows', 'upload CSV ที่ข้อมูลเหมือน DB', 'แสดง rows ที่ SKIP', 'แสดง action=SKIP เพราะไม่มีการเปลี่ยนแปลง', 'Pass', '-'),
    ('TC-015', 'Commit Import', 'กด Commit Import', 'กด commit หลัง preview', 'import สำเร็จ', 'commit สำเร็จ แสดง sync job result', 'Pass', 'สร้าง sync_jobs record'),
    ('TC-016', 'Commit Import', 'ตรวจสอบ ipos table', 'query ipos table หลัง commit', 'data ถูก insert/update', 'ข้อมูลตรงกับ CSV ที่ import, upsert on symbol conflict', 'Pass', '-'),
    ('TC-017', 'Commit Import', 'ตรวจสอบ ipo_financials', 'query ipo_financials หลัง commit', 'financial data ถูก insert', 'proceeds, expense, shares ตรงกับ CSV', 'Pass', 'link ผ่าน ipo_id จาก symbol lookup'),
    ('TC-018', 'Commit Import', 'ตรวจสอบ sectors', 'query sectors หลัง commit', 'sector data ถูก update', 'market/industry/sector mapping ถูกต้อง', 'Pass', '-'),
    ('TC-019', 'Commit Import', 'ตรวจสอบ fa_normalizations', 'query fa_normalizations หลัง commit', 'normalization ถูก update', 'raw_name -> normalized_name mapping ถูกต้อง', 'Pass', 'update fa_companies.normalized_name ด้วย'),
    ('TC-020', 'Commit Import', 'import existing IPO (update)', 'import CSV ที่มี symbol ซ้ำ', 'update record สำเร็จ', 'record ถูก update ด้วยค่าใหม่จาก CSV', 'Pass', 'upsert on conflict(symbol)'),
    ('TC-021', 'Commit Import', 'import IPO ใหม่ (create)', 'import CSV ที่มี symbol ใหม่', 'create record สำเร็จ', 'record ใหม่ถูกสร้างใน ipos table', 'Pass', '-'),
    ('TC-022', 'Auto Status', 'auto-update IPO status', 'import IPO ที่ listing_date ผ่านแล้ว', 'status เป็น matured', 'syncMaturedIpoStatuses() ทำงานหลัง commit', 'Pass', 'เรียก syncMaturedIpoStatuses() อัตโนมัติ'),
    ('TC-023', 'Auto Status', 'auto completeness calculation', 'import IPO ที่มี fields ครบ/ไม่ครบ', 'completeness % ถูกต้อง', 'completeness คำนวณจาก filled fields / total fields', 'Pass', 'แสดงใน v_ipo_completeness view'),
    ('TC-024', 'Auto Status', 'validation_results generated', 'import แล้ว run validation', 'validation records ถูกสร้าง', 'run_validations() ทำงานหลัง commit', 'Pass', 'เรียก POST /api/admin/validation'),
    ('TC-025', 'Preview เปรียบเทียบ', 'preview summary counts', 'ตรวจสอบ counts ใน preview', 'แสดง NEW/UPDATE/SKIP/ERROR counts', 'แสดง summary counts ถูกต้อง', 'Pass', '-'),
    ('TC-026', 'Commit Import', 'commit summary counts', 'ตรวจสอบ counts หลัง commit', 'แสดง inserted/updated/skipped counts', 'counts ตรงกับจำนวน rows ที่ประมวลผล', 'Pass', '-'),
    ('TC-027', 'Edge Cases', 'import large CSV (500+ rows)', 'upload CSV 500+ rows', 'ระบบไม่ crash', 'ระบบประมวลผลสำเร็จ อาจช้าแต่ไม่ timeout', 'Pass', 'ทดสอบกับ base.csv ขนาดใหญ่'),
    ('TC-028', 'Edge Cases', 'import empty CSV', 'upload CSV ที่มีแค่ header', 'validation error', 'แสดง error ว่าไม่มีข้อมูล', 'Pass', '-'),
    ('TC-029', 'Edge Cases', 'import malformed CSV', 'upload CSV ที่ format เสีย', 'parse error', 'แสดง parse error message', 'Pass', 'custom parser จัดการ quoted fields'),
    ('TC-030', 'Edge Cases', 'commit without preview', 'เรียก commit API โดยไม่ preview', 'ระบบ block', 'UI บังคับให้ preview ก่อน, commit button disabled', 'Pass', '-'),
    ('TC-031', 'RBAC & Security', 'unauthorized import (no token)', 'เรียก API โดยไม่มี auth', '401 Unauthorized', 'auth-guard ไม่ครอบคลุม import route ทั้งหมด', 'Fail', 'ต้องเพิ่ม requirePermission() ใน preview route'),
    ('TC-032', 'RBAC & Security', 'readonly role import', 'login ด้วย readonly role', 'permission denied', 'readonly role เข้าถึงหน้า import ได้', 'Fail', 'ต้องเพิ่ม role check ใน AdminNav'),
    ('TC-033', 'RBAC & Security', 'import role (ipos:write)', 'login ด้วย role ที่มี ipos:write', 'import ได้', 'preview และ commit สำเร็จ', 'Pass', "requirePermission('ipos:write')"),
    ('TC-034', 'Audit Log', 'audit log import preview', 'preview แล้วตรวจ audit log', 'มี audit event', 'ยังไม่มี audit event สำหรับ preview', 'N/T', 'audit log ยังไม่ implement สำหรับ preview'),
    ('TC-035', 'Audit Log', 'audit log import commit', 'commit แล้วตรวจ audit log', 'มี audit event', 'ยังไม่มี audit event สำหรับ commit', 'N/T', 'audit log ยังไม่ implement สำหรับ commit'),
    ('TC-036', 'Robustness', 'rollback on DB error', 'simulate DB error ระหว่าง commit', 'partial insert ไม่เกิด', 'ไม่มี explicit transaction wrapping', 'Fail', 'ต้อง wrap ด้วย BEGIN/COMMIT/ROLLBACK'),
    ('TC-037', 'Robustness', 'network interruption', 'simulate network cut ระหว่าง commit', 'transaction safe', 'ยากต่อการจำลอง', 'N/T', 'ต้องทดสอบ manual'),
    ('TC-038', 'Character Encoding', 'UTF-8 Thai characters', 'upload CSV ที่มีภาษาไทย', 'แสดงผลถูกต้อง', 'ภาษาไทยแสดงถูกต้องทั้ง preview และหลัง commit', 'Pass', 'BOM detection ทำงานใน parseCSV()'),
    ('TC-039', 'Character Encoding', 'special characters', 'upload CSV ที่มี quotes, commas', 'parse ถูกต้อง', 'quoted fields parse ถูกต้องตาม RFC 4180', 'Pass', 'custom parser รองรับ'),
    ('TC-040', 'Character Encoding', 'refresh หลัง import', 'import เสร็จแล้ว refresh หน้า', 'data persist', 'ข้อมูลยังอยู่ครบหลัง refresh', 'Pass', 'data อยู่ใน DB ไม่ใช่ client state'),
]

status_map = {'Pass': GREEN_FILL, 'Fail': RED_FILL, 'N/T': YELLOW_FILL}
status_font = {
    'Pass': Font(name='Arial', bold=True, size=10, color='006100'),
    'Fail': Font(name='Arial', bold=True, size=10, color='9C0006'),
    'N/T': Font(name='Arial', bold=True, size=10, color='9C6500'),
}

for ri, tc in enumerate(test_cases, 2):
    row_fill = LIGHT_GRAY if ri % 2 == 0 else WHITE_FILL
    for ci, val in enumerate(tc, 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_WRAP if ci != 7 else CENTER
        cell.border = thin_border
        if ci == 7:
            cell.fill = status_map.get(val, WHITE_FILL)
            cell.font = status_font.get(val, NORMAL_FONT)
        elif ci == 1:
            cell.font = Font(name='Consolas', size=10, bold=True)
            cell.alignment = CENTER
        else:
            cell.fill = row_fill
    ws2.row_dimensions[ri].height = 35

# ===== Sheet 3 =====
ws3 = wb.create_sheet('สรุปปัญหา')
ws3.sheet_properties.tabColor = 'FF0000'

headers3 = ['ลำดับ', 'TC ID', 'ปัญหาที่พบ', 'ความรุนแรง', 'สถานะ', 'แนวทางแก้ไข']
widths3 = [8, 14, 45, 14, 12, 55]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

for ci, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = thin_border

ws3.freeze_panes = 'A2'

sev_fills = {
    'High': PatternFill('solid', fgColor='FFC7CE'),
    'Medium': PatternFill('solid', fgColor='FFEB9C'),
    'Low': PatternFill('solid', fgColor='D6E4F0'),
}

issues = [
    (1, 'TC-031', 'auth-guard ไม่ครอบคลุม import API routes ทั้งหมด (preview route ไม่มี permission check)', 'High', 'Open', 'เพิ่ม requirePermission() ใน preview route.ts และตรวจสอบ commit route.ts ให้ครอบคลุม'),
    (2, 'TC-032', 'readonly role สามารถเข้าถึงหน้า import ได้ เมนู import ไม่ถูกซ่อน', 'Medium', 'Open', 'เพิ่ม role check ใน AdminNav component สำหรับ import menu item, ซ่อนเมนูสำหรับ role ที่ไม่มี ipos:write'),
    (3, 'TC-036', 'commit route ไม่มี explicit transaction wrapping อาจเกิด partial insert เมื่อ DB error', 'High', 'Open', 'wrap upsert operations ทั้งหมดใน commit/route.ts ด้วย BEGIN/COMMIT/ROLLBACK transaction'),
    (4, 'TC-034,\nTC-035', 'audit log ยังไม่ implement สำหรับ import preview และ commit operations', 'Low', 'Open', 'เพิ่ม audit log event สำหรับ import_preview และ import_commit ใน audit_logs table'),
]

for ri, issue in enumerate(issues, 2):
    for ci, val in enumerate(issue, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if ci in (1, 4, 5) else LEFT_WRAP
        cell.border = thin_border
        if ci == 4:
            cell.fill = sev_fills.get(val, WHITE_FILL)
            cell.font = Font(name='Arial', bold=True, size=10)
        if ci == 5:
            cell.fill = RED_FILL
            cell.font = Font(name='Arial', bold=True, size=10, color='9C0006')
    ws3.row_dimensions[ri].height = 45

# ===== Sheet 4 =====
ws4 = wb.create_sheet('ข้อเสนอแนะ')
ws4.sheet_properties.tabColor = 'FFC000'

headers4 = ['ลำดับ', 'ข้อเสนอแนะ', 'Priority', 'TC ที่เกี่ยวข้อง']
widths4 = [8, 65, 14, 18]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

for ci, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = thin_border

ws4.freeze_panes = 'A2'

pri_fills = {
    'สูง': RED_FILL,
    'ปานกลาง': YELLOW_FILL,
    'ต่ำ': SUBHEADER_FILL,
}

recs = [
    (1, 'เพิ่ม transaction wrapping ใน commit API route เพื่อป้องกัน partial insert กรณี DB error ระหว่าง batch upsert', 'สูง', 'TC-036'),
    (2, 'เพิ่ม auth-guard (requirePermission) ให้ครอบคลุม import routes ทั้ง preview และ commit', 'สูง', 'TC-031'),
    (3, 'เพิ่ม RBAC check สำหรับ import menu ใน AdminNav component ซ่อนเมนูสำหรับ readonly role', 'ปานกลาง', 'TC-032'),
    (4, 'เพิ่ม batch size limit เพื่อป้องกัน memory overflow สำหรับไฟล์ CSV ขนาดใหญ่มาก', 'ปานกลาง', 'TC-027'),
    (5, 'implement audit logging สำหรับ import preview และ commit operations', 'ต่ำ', 'TC-034, TC-035'),
    (6, 'เพิ่ม progress indicator (progress bar) สำหรับ large file imports', 'ต่ำ', 'TC-027'),
    (7, 'เพิ่ม CSV template download สำหรับแต่ละประเภท เพื่อให้ user ดาวน์โหลด template ก่อน import', 'ต่ำ', '-'),
    (8, 'เพิ่ม import history page แสดงประวัติ sync_jobs พร้อม status และ error details', 'ต่ำ', '-'),
]

for ri, rec in enumerate(recs, 2):
    for ci, val in enumerate(rec, 1):
        cell = ws4.cell(row=ri, column=ci, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = CENTER if ci in (1, 3, 4) else LEFT_WRAP
        cell.border = thin_border
        if ci == 3:
            cell.fill = pri_fills.get(val, WHITE_FILL)
            cell.font = Font(name='Arial', bold=True, size=10)
    ws4.row_dimensions[ri].height = 35

output = r'D:\IPO\ipo-ui\reports\import-csv-test-report.xlsx'
wb.save(output)
print(f'Saved to {output}')
